from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta, date
import os
from functools import wraps
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io
from config import config
import time

# Load environment variables
load_dotenv()

# Create Flask app
app = Flask(__name__)

# Load configuration based on environment
config_name = os.environ.get('FLASK_CONFIG', 'development')
app.config.from_object(config[config_name])

# Ensure instance directory exists for SQLite
if 'sqlite' in app.config.get('SQLALCHEMY_DATABASE_URI', ''):
    db_path = app.config['SQLALCHEMY_DATABASE_URI'].replace('sqlite:///', '')
    db_dir = os.path.dirname(db_path)
    if db_dir and not os.path.exists(db_dir):
        os.makedirs(db_dir, exist_ok=True)

# Initialize SQLAlchemy
db = SQLAlchemy(app)

# Association table for many-to-many relationship between students and groups
student_groups = db.Table('student_groups',
    db.Column('student_id', db.Integer, db.ForeignKey('student.id'), primary_key=True),
    db.Column('group_id', db.Integer, db.ForeignKey('group.id'), primary_key=True)
)

# Database Models
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(128), nullable=False)
    full_name = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(20), default='instructor')  # 'admin' or 'instructor'
    instructor_id = db.Column(db.Integer, db.ForeignKey('instructor.id'), nullable=True)  # Link to instructor if role is instructor
    is_hidden = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    last_login = db.Column(db.DateTime)
    last_activity = db.Column(db.DateTime, default=datetime.utcnow)
    is_online = db.Column(db.Boolean, default=False)
    
    # Relationship
    linked_instructor = db.relationship('Instructor', backref='user_account', uselist=False)
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)
    
    def update_activity(self):
        """Update last activity and set user as online"""
        self.last_activity = datetime.utcnow()
        self.is_online = True
        db.session.commit()
    
    def is_active_now(self):
        """Check if user is active (last activity within 5 minutes)"""
        if not self.last_activity:
            return False
        return (datetime.utcnow() - self.last_activity).total_seconds() < 300  # 5 minutes

class Instructor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    phone = db.Column(db.String(20))
    specialization = db.Column(db.String(100))
    students = db.relationship('Student', backref='instructor_ref', lazy=True)
    groups = db.relationship('Group', backref='instructor_ref', lazy=True)

class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    phone = db.Column(db.String(20))
    age = db.Column(db.Integer)
    location = db.Column(db.String(50))  # Changed from level to location
    instructor_id = db.Column(db.Integer, db.ForeignKey('instructor.id'))
    # Removed group_id - now using many-to-many relationship
    total_paid = db.Column(db.Float, default=0.0)
    discount = db.Column(db.Float, default=0.0)  # Discount amount in currency
    # Removed course_price - now price is per group
    registration_date = db.Column(db.DateTime, nullable=False)
    # Many-to-many relationship with groups
    groups = db.relationship('Group', secondary=student_groups, backref=db.backref('students', lazy='dynamic'))
    
    @property
    def total_course_price(self):
        """Calculate total price of all groups the student is enrolled in"""
        return sum(group.price for group in self.groups)
    
    @property
    def total_course_price_after_discount(self):
        """Calculate total price after applying discount"""
        total_price = self.total_course_price
        discounted_price = total_price - self.discount
        return max(0, discounted_price)  # Ensure price doesn't go below 0

    @property
    def remaining_balance(self):
        """Calculate remaining balance for the student after discount"""
        return self.total_course_price_after_discount - self.total_paid

class Group(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    level = db.Column(db.String(50))
    instructor_id = db.Column(db.Integer, db.ForeignKey('instructor.id'))
    max_students = db.Column(db.Integer, default=15)
    price = db.Column(db.Float, default=0.0)  # Price for this group
    # Students relationship is now defined in Student model with secondary table
    schedules = db.relationship('Schedule', backref='group_ref', lazy=True)

class Schedule(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    group_id = db.Column(db.Integer, db.ForeignKey('group.id'))
    day_of_week = db.Column(db.String(20))  # السبت، الأحد، الاثنين، etc.
    start_time = db.Column(db.String(10))
    end_time = db.Column(db.String(10))

class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'))
    date = db.Column(db.Date)
    status = db.Column(db.String(20))  # حاضر، غائب، متأخر
    group_id = db.Column(db.Integer, db.ForeignKey('group.id'))

class Payment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'))
    amount = db.Column(db.Float)
    date = db.Column(db.DateTime, default=datetime.utcnow)
    month = db.Column(db.String(20))
    notes = db.Column(db.Text)

class Expense(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    description = db.Column(db.String(200), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    category = db.Column(db.String(100))  # رواتب، إيجار، مرافق، مستلزمات، أخرى
    date = db.Column(db.DateTime, default=datetime.utcnow)
    notes = db.Column(db.Text)

class Task(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    priority = db.Column(db.String(20), default='متوسط')  # عالي، متوسط، منخفض
    status = db.Column(db.String(20), default='قيد التنفيذ')  # قيد التنفيذ، مكتمل، ملغي
    due_date = db.Column(db.Date)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    completed_at = db.Column(db.DateTime)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    assigned_to = db.Column(db.Integer, db.ForeignKey('user.id'))
    
    # Relationships
    creator = db.relationship('User', foreign_keys=[created_by], backref='created_tasks')
    assignee = db.relationship('User', foreign_keys=[assigned_to], backref='assigned_tasks')

    @property
    def is_overdue(self):
        """Check if task is overdue"""
        if self.due_date and self.status != 'مكتمل':
            return datetime.now().date() > self.due_date
        return False

    @property
    def days_remaining(self):
        """Calculate days remaining until due date"""
        if self.due_date and self.status != 'مكتمل':
            delta = self.due_date - datetime.now().date()
            return delta.days
        return None

class Note(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    content = db.Column(db.Text, nullable=False)
    category = db.Column(db.String(50), default='عام')  # عام، شخصي، عمل، مهم
    color = db.Column(db.String(20), default='yellow')  # yellow, blue, green, red, purple
    is_pinned = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    
    # Relationship
    creator = db.relationship('User', backref='notes')

    @property
    def created_ago(self):
        """Get how long ago the note was created"""
        delta = datetime.utcnow() - self.created_at
        if delta.days > 0:
            return f'منذ {delta.days} يوم'
        elif delta.seconds > 3600:
            return f'منذ {delta.seconds // 3600} ساعة'
        elif delta.seconds > 60:
            return f'منذ {delta.seconds // 60} دقيقة'
        else:
            return 'منذ لحظات'

class InstructorNote(db.Model):
    """Notes created by instructors - automatically sent to admins"""
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    content = db.Column(db.Text, nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=True)  # Optional: specific student
    group_id = db.Column(db.Integer, db.ForeignKey('group.id'), nullable=True)  # Optional: specific group
    priority = db.Column(db.String(20), default='متوسط')  # عالي، متوسط، منخفض
    status = db.Column(db.String(20), default='جديد')  # جديد، قيد المراجعة، مكتمل
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))  # Instructor user
    reviewed_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)  # Admin who reviewed
    reviewed_at = db.Column(db.DateTime, nullable=True)
    admin_response = db.Column(db.Text, nullable=True)
    
    # Relationships
    creator = db.relationship('User', foreign_keys=[created_by], backref='instructor_notes')
    reviewer = db.relationship('User', foreign_keys=[reviewed_by])
    student = db.relationship('Student', backref='instructor_notes')
    group = db.relationship('Group', backref='instructor_notes')

    @property
    def created_ago(self):
        """Get how long ago the note was created"""
        delta = datetime.utcnow() - self.created_at
        if delta.days > 0:
            return f'منذ {delta.days} يوم'
        elif delta.seconds > 3600:
            return f'منذ {delta.seconds // 3600} ساعة'
        elif delta.seconds > 60:
            return f'منذ {delta.seconds // 60} دقيقة'
        else:
            return 'منذ لحظات'

class InstructorTodo(db.Model):
    """Todo list for instructors - personal task management"""
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    status = db.Column(db.String(20), default='مفتوح')  # مفتوح، مكتمل، ملغي
    priority = db.Column(db.String(20), default='متوسط')  # عالي، متوسط، منخفض
    category = db.Column(db.String(50), default='عام')  # عام، تحضير، حضور، متابعة، إداري
    group_id = db.Column(db.Integer, db.ForeignKey('group.id'), nullable=True)  # Optional: specific group
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=True)  # Optional: specific student
    due_date = db.Column(db.Date, nullable=True)  # Optional: due date
    reminder_date = db.Column(db.DateTime, nullable=True)  # Optional: reminder date
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    completed_at = db.Column(db.DateTime, nullable=True)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))  # Instructor user
    
    # Relationships
    creator = db.relationship('User', backref='instructor_todos')
    group = db.relationship('Group', backref='instructor_todos')
    student = db.relationship('Student', backref='instructor_todos')

    @property
    def is_overdue(self):
        """Check if todo is overdue"""
        if self.due_date and self.status == 'مفتوح':
            return datetime.now().date() > self.due_date
        return False

    @property
    def days_remaining(self):
        """Calculate days remaining until due date"""
        if self.due_date and self.status == 'مفتوح':
            delta = self.due_date - datetime.now().date()
            return delta.days
        return None

    @property
    def created_ago(self):
        """Get how long ago the todo was created"""
        delta = datetime.utcnow() - self.created_at
        if delta.days > 0:
            return f'منذ {delta.days} يوم'
        elif delta.seconds > 3600:
            return f'منذ {delta.seconds // 3600} ساعة'
        elif delta.seconds > 60:
            return f'منذ {delta.seconds // 60} دقيقة'
        else:
            return 'منذ لحظات'

# Update user activity before each request
@app.before_request
def update_user_activity():
    if 'user_id' in session:
        user = User.query.get(session['user_id'])
        if user:
            user.update_activity()

# Authentication functions
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('يرجى تسجيل الدخول للوصول إلى هذه الصفحة', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('يجب تسجيل الدخول أولاً', 'error')
            return redirect(url_for('login'))
        
        user = User.query.get(session['user_id'])
        if not user or user.role != 'admin':
            flash('ليس لديك صلاحية للوصول لهذه الصفحة', 'error')
            return redirect(url_for('index'))
        
        return f(*args, **kwargs)
    return decorated_function

def instructor_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('يجب تسجيل الدخول أولاً', 'error')
            return redirect(url_for('login'))
        
        user = User.query.get(session['user_id'])
        if not user or user.role not in ['admin', 'instructor']:
            flash('ليس لديك صلاحية للوصول لهذه الصفحة', 'error')
            return redirect(url_for('index'))
        
        return f(*args, **kwargs)
    return decorated_function

def get_current_user():
    if 'user_id' in session:
        return User.query.get(session['user_id'])
    return None

def create_default_admin():
    """Create default hidden admin user if it doesn't exist"""
    admin_user = User.query.filter_by(username='araby').first()
    if not admin_user:
        admin_user = User(
            username='araby',
            full_name='System Administrator',
            role='admin',
            is_hidden=True
        )
        admin_user.set_password('92321066')
        db.session.add(admin_user)
        db.session.commit()
        print("Default admin user 'araby' created successfully!")

# Helper function to get Arabic day name
def get_arabic_day_name(date_obj):
    arabic_days = {
        'Monday': 'الاثنين',
        'Tuesday': 'الثلاثاء', 
        'Wednesday': 'الأربعاء',
        'Thursday': 'الخميس',
        'Friday': 'الجمعة',
        'Saturday': 'السبت',
        'Sunday': 'الأحد'
    }
    english_day = date_obj.strftime('%A')
    return arabic_days.get(english_day, english_day)

# Helper function to format date in Arabic
def format_arabic_date(date_obj):
    """Format date as: 4 يونيو 2025 (Day Month Year)"""
    if not date_obj:
        return ""
    
    arabic_months = {
        1: 'يناير', 2: 'فبراير', 3: 'مارس', 4: 'أبريل',
        5: 'مايو', 6: 'يونيو', 7: 'يوليو', 8: 'أغسطس',
        9: 'سبتمبر', 10: 'أكتوبر', 11: 'نوفمبر', 12: 'ديسمبر'
    }
    
    day = date_obj.day
    month = arabic_months[date_obj.month]
    year = date_obj.year
    
    # Format: Day Month Year (e.g., 4 يونيو 2025)
    return f"{day} {month} {year}"

def format_time_12hour(datetime_obj):
    """Format time in 12-hour format with Arabic AM/PM"""
    if not datetime_obj:
        return ""
    
    time_str = datetime_obj.strftime('%I:%M')
    am_pm = datetime_obj.strftime('%p')
    
    # Convert AM/PM to Arabic
    if am_pm == 'AM':
        am_pm_arabic = 'ص'
    else:
        am_pm_arabic = 'م'
    
    return f"{time_str} {am_pm_arabic}"

def format_date_for_input(date_obj):
    """Format date for HTML input fields as DD-MM-YYYY"""
    if not date_obj:
        return ""
    
    return date_obj.strftime('%d-%m-%Y')

def parse_date_from_input(date_string):
    """Parse date from HTML5 date input (YYYY-MM-DD) or DD-MM-YYYY format to datetime object"""
    if not date_string:
        return None
    
    try:
        # Try YYYY-MM-DD format first (HTML5 date input format)
        return datetime.strptime(date_string, '%Y-%m-%d')
    except ValueError:
        try:
            # Fallback to DD-MM-YYYY format (manual input)
            return datetime.strptime(date_string, '%d-%m-%Y')
        except ValueError:
            # If both fail, return None
            return None

# Function to get today's schedule
def get_today_schedule():
    today = datetime.now()
    today_arabic = get_arabic_day_name(today)
    
    # Get all schedules for today
    today_schedules = Schedule.query.filter_by(day_of_week=today_arabic).all()
    
    schedule_data = []
    for schedule in today_schedules:
        group = Group.query.get(schedule.group_id)
        if group and group.instructor_ref:
            schedule_data.append({
                'group_name': group.name,
                'instructor_name': group.instructor_ref.name,
                'start_time': schedule.start_time,
                'end_time': schedule.end_time,
                'level': group.level,
                'student_count': group.students.count(),  # Use count() for dynamic relationship
                'max_students': group.max_students  # Add max_students field
            })
    
    # Sort by start time
    schedule_data.sort(key=lambda x: x['start_time'])
    return schedule_data

# Function to get weekly schedule
def get_weekly_schedule():
    """Get schedule for all days of the week"""
    arabic_days = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
    weekly_schedule = {}
    
    for day in arabic_days:
        day_schedules = Schedule.query.filter_by(day_of_week=day).all()
        schedule_data = []
        
        for schedule in day_schedules:
            group = Group.query.get(schedule.group_id)
            if group and group.instructor_ref:
                schedule_data.append({
                    'group_name': group.name,
                    'instructor_name': group.instructor_ref.name,
                    'start_time': schedule.start_time,
                    'end_time': schedule.end_time,
                    'level': group.level,
                    'student_count': group.students.count(),  # Use count() for dynamic relationship
                    'max_students': group.max_students,  # Add max_students field
                    'group_id': group.id
                })
        
        # Sort by start time
        schedule_data.sort(key=lambda x: x['start_time'])
        weekly_schedule[day] = schedule_data
    
    return weekly_schedule

# Routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        remember_me = 'remember_me' in request.form
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            session['user_id'] = user.id
            session['username'] = user.username
            session['user_role'] = user.role
            session['user_name'] = user.full_name
            
            # Set session as permanent if remember me is checked
            if remember_me:
                session.permanent = True
                app.permanent_session_lifetime = timedelta(days=30)  # Remember for 30 days
            
            # Update last login and activity
            user.last_login = datetime.utcnow()
            user.update_activity()
            
            flash(f'مرحباً {user.full_name}!', 'success')
            return redirect(url_for('index'))
        else:
            flash('اسم المستخدم أو كلمة المرور غير صحيحة', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('تم تسجيل الخروج بنجاح! نراك قريباً 👋', 'success')
    return redirect(url_for('login'))

@app.route('/users')
@admin_required
def users():
    users = User.query.filter_by(is_hidden=False).all()
    instructors = Instructor.query.all()
    current_user = get_current_user()
    return render_template('users.html', users=users, instructors=instructors, current_user=current_user)

@app.route('/add_user', methods=['POST'])
@admin_required
def add_user():
    username = request.form['username']
    password = request.form['password']
    full_name = request.form['full_name']
    role = request.form['role']
    instructor_id = request.form.get('instructor_id') if role == 'instructor' else None
    
    # Check if username already exists
    if User.query.filter_by(username=username).first():
        flash('اسم المستخدم موجود بالفعل', 'error')
        return redirect(url_for('users'))
    
    # If instructor role, check if instructor is already linked to another user
    if role == 'instructor' and instructor_id:
        existing_user = User.query.filter_by(instructor_id=instructor_id).first()
        if existing_user:
            flash('هذا المدرس مرتبط بمستخدم آخر بالفعل', 'error')
            return redirect(url_for('users'))
    
    new_user = User(
        username=username,
        full_name=full_name,
        role=role,
        instructor_id=int(instructor_id) if instructor_id else None
    )
    new_user.set_password(password)
    
    try:
        db.session.add(new_user)
        db.session.commit()
        flash('تم إضافة المستخدم بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash('حدث خطأ أثناء إضافة المستخدم', 'error')
    
    return redirect(url_for('users'))

@app.route('/edit_user/<int:user_id>', methods=['POST'])
@admin_required
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    current_user = get_current_user()
    
    # Prevent editing hidden admin unless you are the hidden admin
    if user.is_hidden and current_user.username != 'araby':
        flash('لا يمكن تعديل هذا المستخدم', 'error')
        return redirect(url_for('users'))
    
    user.username = request.form['username']
    user.full_name = request.form['full_name']
    user.role = request.form['role']
    
    if request.form['password']:
        user.set_password(request.form['password'])
    
    db.session.commit()
    flash('تم تحديث بيانات المستخدم بنجاح', 'success')
    return redirect(url_for('users'))

@app.route('/delete_user/<int:user_id>', methods=['POST'])
@admin_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    current_user = get_current_user()
    
    # Prevent deleting hidden admin unless you are the hidden admin
    if user.is_hidden and current_user.username != 'araby':
        flash('لا يمكن حذف هذا المستخدم', 'error')
        return redirect(url_for('users'))
    
    # Prevent users from deleting themselves
    if user.id == current_user.id:
        flash('لا يمكن حذف حسابك الشخصي', 'error')
        return redirect(url_for('users'))
    
    db.session.delete(user)
    db.session.commit()
    flash('تم حذف المستخدم بنجاح', 'success')
    return redirect(url_for('users'))

@app.route('/')
@login_required
def index():
    current_user = get_current_user()
    
    if current_user.role == 'instructor':
        return redirect(url_for('instructor_dashboard'))
    
    # Original admin/user dashboard code
    students = Student.query.all()
    instructors = Instructor.query.all() 
    groups = Group.query.all()
    
    total_students = len(students)
    total_groups = len(groups)
    total_instructors = len(instructors)
    
    # Get today's schedule
    today_schedule = get_today_schedule()
    
    # Get weekly schedule  
    weekly_schedule = get_weekly_schedule()
    
    # Get today's Arabic day name
    today_arabic = get_arabic_day_name(datetime.now())
    
    return render_template('index.html', 
                         total_students=total_students,
                         total_groups=total_groups, 
                         total_instructors=total_instructors,
                         today_schedule=today_schedule,
                         weekly_schedule=weekly_schedule,
                         today_date=datetime.now(),
                         today_arabic=today_arabic)

@app.route('/instructor_dashboard')
@instructor_required
def instructor_dashboard():
    current_user = get_current_user()
    
    if not current_user.linked_instructor:
        flash('حساب المدرس غير مرتبط بملف مدرس', 'error')
        return redirect(url_for('logout'))
    
    instructor = current_user.linked_instructor
    
    # Get instructor's groups and students
    instructor_groups = instructor.groups
    instructor_students = get_instructor_students(current_user)
    
    # Get today's schedule for this instructor
    today_schedule = []
    for group in instructor_groups:
        for schedule in group.schedules:
            if schedule.day_of_week == get_arabic_day_name(datetime.now()):
                today_schedule.append({
                    'group': group,
                    'schedule': schedule
                })
    
    # Get recent instructor notes
    recent_notes = InstructorNote.query.filter_by(created_by=current_user.id)\
                                      .order_by(InstructorNote.created_at.desc())\
                                      .limit(5).all()
    
    # Statistics
    total_students = len(instructor_students)
    total_groups = len(instructor_groups)
    
    # Get attendance statistics for instructor's groups
    total_classes_today = len(today_schedule)
    
    # Get unique ages for instructor's students
    instructor_ages = []
    for student in instructor_students:
        if student.age and student.age not in instructor_ages:
            instructor_ages.append(student.age)
    instructor_ages.sort()
    
    return render_template('instructor_dashboard.html',
                         instructor=instructor,
                         total_students=total_students,
                         total_groups=total_groups,
                         total_classes_today=total_classes_today,
                         today_schedule=today_schedule,
                         recent_notes=recent_notes,
                         instructor_groups=instructor_groups,
                         instructor_students=instructor_students,
                         instructor_ages=instructor_ages)

@app.route('/students')
@login_required
def students():
    # Get filter parameters
    group_filter = request.args.get('group_id', '')
    age_filter = request.args.get('age_range', '')
    location_filter = request.args.get('location', '')
    
    # Start with base query
    query = Student.query
    
    # Apply group filter - since students can have multiple groups, we need to join
    if group_filter:
        query = query.join(Student.groups).filter(Group.id == int(group_filter))
    
    # Apply age filter
    if age_filter:
        try:
            age_value = int(age_filter)
            query = query.filter(Student.age == age_value)
        except ValueError:
            pass  # Ignore invalid age values
    
    # Apply location filter
    if location_filter:
        query = query.filter(Student.location.ilike(f'%{location_filter}%'))
    
    students = query.all()
    instructors = Instructor.query.all()
    groups = Group.query.all()
    
    # Get all unique locations for the filter dropdown
    locations = db.session.query(Student.location).filter(Student.location.isnot(None)).distinct().all()
    locations = [loc[0] for loc in locations if loc[0] and loc[0].strip()]
    locations.sort()
    
    # Get all unique ages for the filter dropdown
    ages = db.session.query(Student.age).filter(Student.age.isnot(None)).distinct().all()
    ages = [age[0] for age in ages if age[0] is not None]
    ages.sort()
    
    return render_template('students.html', 
                         students=students, 
                         instructors=instructors, 
                         groups=groups,
                         locations=locations,
                         ages=ages,
                         selected_group=group_filter,
                         selected_age=age_filter,
                         selected_location=location_filter)

@app.route('/add_student', methods=['POST'])
def add_student():
    try:
        name = request.form['name']
        phone = request.form.get('phone', '')  # Use get() for optional fields
        age = int(request.form['age'])
        location = request.form.get('location', '')  # Use get() for optional fields
        instructor_id = None  # Make instructor optional - can be set later if needed
        registration_date = parse_date_from_input(request.form['registration_date'])
        discount = float(request.form.get('discount', 0))  # Get discount amount, default to 0
        
        if not registration_date:
            flash('يرجى اختيار تاريخ التسجيل من منتقي التاريخ', 'error')
            return redirect(url_for('students'))
        
        student = Student(
            name=name,
            phone=phone,
            age=age,
            location=location,
            instructor_id=instructor_id,
            registration_date=registration_date,
            discount=discount
        )
        
        db.session.add(student)
        db.session.commit()
        
        # Handle group selections (multiple groups allowed)
        group_ids = request.form.getlist('group_ids')  # Get list of selected group IDs
        if group_ids:
            for group_id in group_ids:
                if group_id:  # Make sure it's not empty
                    group = Group.query.get(int(group_id))
                    if group:
                        student.groups.append(group)
        
        db.session.commit()
        flash('تم إضافة الطالب بنجاح!', 'success')
        return redirect(url_for('students'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء إضافة الطالب: {str(e)}', 'error')
        return redirect(url_for('students'))

@app.route('/instructors')
@login_required
def instructors():
    instructors = Instructor.query.all()
    return render_template('instructors.html', instructors=instructors)

@app.route('/add_instructor', methods=['POST'])
def add_instructor():
    name = request.form['name']
    phone = request.form['phone']
    specialization = request.form['specialization']
    
    instructor = Instructor(
        name=name,
        phone=phone,
        specialization=specialization
    )
    
    db.session.add(instructor)
    db.session.commit()
    flash('تم إضافة المدرس بنجاح', 'success')
    return redirect(url_for('instructors'))

@app.route('/groups')
@login_required
def groups():
    # Get filter parameters
    instructor_filter = request.args.get('instructor_id', type=int)
    
    # Start with base query
    groups_query = Group.query
    
    # Apply instructor filter if specified
    if instructor_filter:
        groups_query = groups_query.filter(Group.instructor_id == instructor_filter)
    
    groups = groups_query.all()
    instructors = Instructor.query.all()
    
    # Calculate total students across filtered groups
    total_students = 0
    for group in groups:
        total_students += group.students.count()
    
    return render_template('groups.html', 
                         groups=groups, 
                         instructors=instructors,
                         total_students=total_students,
                         selected_instructor=instructor_filter)

def check_instructor_schedule_conflicts(day, start_time, end_time, instructor_id, exclude_group_id=None):
    """Check for schedule conflicts for the same instructor only"""
    conflicts = []
    
    # Get all schedules for the same day and instructor
    existing_schedules = db.session.query(Schedule).join(Group).filter(
        Schedule.day_of_week == day,
        Group.instructor_id == instructor_id
    )
    
    # Exclude current group if editing
    if exclude_group_id:
        existing_schedules = existing_schedules.filter(Group.id != exclude_group_id)
    
    existing_schedules = existing_schedules.all()
    
    # Convert times to minutes for easier comparison
    def time_to_minutes(time_str):
        hours, minutes = map(int, time_str.split(':'))
        return hours * 60 + minutes
    
    new_start_min = time_to_minutes(start_time)
    new_end_min = time_to_minutes(end_time)
    
    for schedule in existing_schedules:
        existing_start_min = time_to_minutes(schedule.start_time)
        existing_end_min = time_to_minutes(schedule.end_time)
        
        # Check for overlap
        if (new_start_min < existing_end_min and new_end_min > existing_start_min):
            conflicts.append({
                'group_name': schedule.group_ref.name,
                'start_time': schedule.start_time,
                'end_time': schedule.end_time,
                'day': day
            })
    
    return conflicts



@app.route('/add_group', methods=['POST'])
def add_group():
    name = request.form['name']
    level = request.form['level']
    instructor_id = int(request.form['instructor_id'])
    max_students = int(request.form['max_students'])
    price = float(request.form['price'])
    force_save = request.form.get('force_save', 'false') == 'true'
    
    # Collect schedule data for conflict checking
    selected_days = request.form.getlist('days[]')
    schedules_to_add = []
    
    for day in selected_days:
        day_prefix = {
            'السبت': 'sat',
            'الأحد': 'sun', 
            'الاثنين': 'mon',
            'الثلاثاء': 'tue',
            'الأربعاء': 'wed',
            'الخميس': 'thu',
            'الجمعة': 'fri'
        }.get(day)
        
        if not day_prefix:
            continue
            
        hour = request.form.get(f'{day_prefix}_hour')
        minute = request.form.get(f'{day_prefix}_minute')
        period = request.form.get(f'{day_prefix}_period')
        duration = request.form.get(f'{day_prefix}_duration')
        
        if hour and minute and period and duration:
            start_time = convert_12_to_24_hour(hour, minute, period)
            
            # Calculate end time based on duration
            duration_minutes = int(duration)
            start_total_minutes = int(start_time.split(':')[0]) * 60 + int(start_time.split(':')[1])
            end_total_minutes = start_total_minutes + duration_minutes
            end_hour = (end_total_minutes // 60) % 24
            end_minute = end_total_minutes % 60
            end_time = f"{end_hour:02d}:{end_minute:02d}"
            
            schedules_to_add.append({
                'day': day,
                'start_time': start_time,
                'end_time': end_time
            })
    
    # Check for instructor schedule conflicts if not forcing save
    all_conflicts = []
    if not force_save and schedules_to_add:
        for schedule_data in schedules_to_add:
            conflicts = check_instructor_schedule_conflicts(
                schedule_data['day'], 
                schedule_data['start_time'], 
                schedule_data['end_time'], 
                instructor_id
            )
            all_conflicts.extend(conflicts)
        
        if all_conflicts:
            # Get instructor name
            instructor = Instructor.query.get(instructor_id)
            instructor_name = instructor.name if instructor else "غير محدد"
            
            # Return conflict information to frontend
            conflict_message = f"المدرس <strong>{instructor_name}</strong> لديه مجموعة أخرى في نفس التوقيت:<br>"
            for conflict in all_conflicts:
                start_12 = convert_24_to_12_hour(conflict['start_time'])
                end_12 = convert_24_to_12_hour(conflict['end_time'])
                conflict_message += f"• مجموعة {conflict['group_name']} - {conflict['day']}: {start_12['hour']}:{start_12['minute']} {start_12['period']} - {end_12['hour']}:{end_12['minute']} {end_12['period']}<br>"
            
            return jsonify({
                'has_conflicts': True,
                'message': conflict_message,
                'form_data': dict(request.form)
            })
    
    # Create and save group
    group = Group(
        name=name,
        level=level,
        instructor_id=instructor_id,
        max_students=max_students,
        price=price
    )
    
    db.session.add(group)
    db.session.commit()
    
    # Add schedules
    for schedule_data in schedules_to_add:
        schedule = Schedule(
            group_id=group.id,
            day_of_week=schedule_data['day'],
            start_time=schedule_data['start_time'],
            end_time=schedule_data['end_time']
        )
        db.session.add(schedule)
    
    db.session.commit()
    flash('تم إضافة المجموعة بنجاح', 'success')
    
    # Check if this is an AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': True, 'redirect': url_for('groups')})
    return redirect(url_for('groups'))

@app.route('/attendance')
@login_required
def attendance():
    groups = Group.query.all()
    students = Student.query.all()
    today = datetime.now().date()
    return render_template('attendance.html', groups=groups, students=students, today=today)

@app.route('/mark_attendance', methods=['POST'])
def mark_attendance():
    data = request.get_json()
    date = datetime.strptime(data['date'], '%Y-%m-%d').date()
    group_id = data['group_id']
    
    for student_data in data['students']:
        student_id = student_data['student_id']
        status = student_data['status']
        
        # Check if attendance already exists
        existing = Attendance.query.filter_by(
            student_id=student_id,
            date=date,
            group_id=group_id
        ).first()
        
        if existing:
            existing.status = status
        else:
            attendance = Attendance(
                student_id=student_id,
                date=date,
                status=status,
                group_id=group_id
            )
            db.session.add(attendance)
    
    db.session.commit()
    return jsonify({'success': True, 'message': 'تم حفظ الحضور بنجاح'})

@app.route('/payments')
@login_required
def payments():
    students = Student.query.all()
    payments = Payment.query.order_by(Payment.date.desc()).all()
    expenses = Expense.query.order_by(Expense.date.desc()).all()
    
    # Calculate comprehensive statistics
    total_income = sum(payment.amount for payment in payments) if payments else 0
    total_expenses = sum(expense.amount for expense in expenses) if expenses else 0
    net_balance = total_income - total_expenses
    
    students_with_dues = sum(1 for student in students if student.remaining_balance > 0)
    recent_payments = len([p for p in payments if (datetime.now() - p.date).days <= 30]) if payments else 0
    recent_expenses = len([e for e in expenses if (datetime.now() - e.date).days <= 30]) if expenses else 0
    
    # Monthly breakdown for current year
    current_year = datetime.now().year
    monthly_income = {}
    monthly_expenses = {}
    
    # Get monthly income
    for payment in payments:
        if payment.date.year == current_year:
            month = payment.date.month
            monthly_income[month] = monthly_income.get(month, 0) + payment.amount
    
    # Get monthly expenses  
    for expense in expenses:
        if expense.date.year == current_year:
            month = expense.date.month
            monthly_expenses[month] = monthly_expenses.get(month, 0) + expense.amount
    
    return render_template('payments.html', 
                         students=students, 
                         payments=payments,
                         expenses=expenses,
                         total_income=total_income,
                         total_expenses=total_expenses,
                         net_balance=net_balance,
                         students_with_dues=students_with_dues,
                         recent_payments=recent_payments,
                         recent_expenses=recent_expenses,
                         monthly_income=monthly_income,
                         monthly_expenses=monthly_expenses)

@app.route('/add_payment', methods=['POST'])
def add_payment():
    student_id = int(request.form['student_id'])
    amount = float(request.form['amount'])
    month = request.form['month']
    notes = request.form['notes']
    
    payment = Payment(
        student_id=student_id,
        amount=amount,
        month=month,
        notes=notes
    )
    
    # Update student's total paid
    student = Student.query.get(student_id)
    student.total_paid += amount
    
    db.session.add(payment)
    db.session.commit()
    flash('تم إضافة الدفعة بنجاح', 'success')
    return redirect(url_for('payments'))

@app.route('/reports')
@login_required
def reports():
    # Attendance statistics
    total_students = Student.query.count()
    today = datetime.now().date()
    present_today = Attendance.query.filter_by(date=today, status='حاضر').count()
    absent_today = Attendance.query.filter_by(date=today, status='غائب').count()
    
    # Payment statistics
    total_revenue = db.session.query(db.func.sum(Payment.amount)).scalar() or 0
    
    # Calculate pending payments based on group-based pricing after discounts
    pending_payments = 0
    students = Student.query.all()
    for student in students:
        if student.remaining_balance > 0:
            pending_payments += student.remaining_balance
    
    # Other statistics
    groups_count = Group.query.count()
    instructors_count = Instructor.query.count()
    today_date = datetime.now().strftime('%Y-%m-%d')
    
    # Additional useful statistics - calculate expected revenue after discounts
    total_groups_revenue = sum(student.total_course_price_after_discount for student in Student.query.all())
    late_today = Attendance.query.filter_by(date=today, status='متأخر').count()
    
    # Monthly statistics for the current year
    current_year = datetime.now().year
    monthly_payments = {}
    monthly_expenses = {}
    
    # Get monthly payment data
    payments = Payment.query.filter(
        db.extract('year', Payment.date) == current_year
    ).all()
    
    for payment in payments:
        month = payment.date.month
        monthly_payments[month] = monthly_payments.get(month, 0) + payment.amount
    
    # Get monthly expense data
    expenses = Expense.query.filter(
        db.extract('year', Expense.date) == current_year
    ).all()
    
    for expense in expenses:
        month = expense.date.month
        monthly_expenses[month] = monthly_expenses.get(month, 0) + expense.amount
    
    return render_template('reports.html',
                         total_students=total_students,
                         present_today=present_today,
                         absent_today=absent_today,
                         late_today=late_today,
                         total_revenue=total_revenue,
                         pending_payments=pending_payments,
                         groups_count=groups_count,
                         instructors_count=instructors_count,
                         today_date=datetime.now(),  # Pass datetime object instead of string
                         total_groups_revenue=total_groups_revenue,
                         monthly_payments=monthly_payments,
                         monthly_expenses=monthly_expenses)

@app.route('/export_reports')
@login_required
def export_reports():
    """Export comprehensive reports to Excel file"""
    try:
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير شامل"
        
        # Set RTL direction for Arabic support
        ws.sheet_view.rightToLeft = True
        
        # Define styles
        header_font = Font(size=14, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        sub_header_font = Font(size=12, bold=True, color="2F5F8F")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        current_row = 1
        
        # Title
        ws.merge_cells(f'A{current_row}:F{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = f"تقرير شامل - مركز تفرا التعليمي - {format_arabic_date(datetime.now())}"
        title_cell.font = Font(size=16, bold=True, color="2F5F8F")
        title_cell.alignment = center_alignment
        current_row += 2
        
        # Basic Statistics Section
        ws[f'A{current_row}'] = "الإحصائيات الأساسية"
        ws[f'A{current_row}'].font = sub_header_font
        current_row += 1
        
        # Get statistics data
        total_students = Student.query.count()
        instructors_count = Instructor.query.count()
        groups_count = Group.query.count()
        today = datetime.now().date()
        present_today = Attendance.query.filter_by(date=today, status='حاضر').count()
        absent_today = Attendance.query.filter_by(date=today, status='غائب').count()
        late_today = Attendance.query.filter_by(date=today, status='متأخر').count()
        
        # Add basic statistics
        stats_data = [
            ['البيان', 'القيمة'],
            ['إجمالي الطلاب', total_students],
            ['عدد المدرسين', instructors_count],
            ['عدد المجموعات', groups_count],
            ['حاضر اليوم', present_today],
            ['غائب اليوم', absent_today],
            ['متأخر اليوم', late_today],
        ]
        
        for row_data in stats_data:
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
                if current_row == len(stats_data) + current_row - len(stats_data):  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
            current_row += 1
        
        current_row += 2
        
        # Financial Statistics Section
        ws[f'A{current_row}'] = "الإحصائيات المالية"
        ws[f'A{current_row}'].font = sub_header_font
        current_row += 1
        
        # Get financial data
        total_revenue = db.session.query(db.func.sum(Payment.amount)).scalar() or 0
        total_expenses = db.session.query(db.func.sum(Expense.amount)).scalar() or 0
        pending_payments = sum(student.remaining_balance for student in Student.query.all() if student.remaining_balance > 0)
        
        financial_data = [
            ['البيان المالي', 'المبلغ (ريال)'],
            ['إجمالي الإيرادات', f"{total_revenue:,.0f}"],
            ['إجمالي المصروفات', f"{total_expenses:,.0f}"],
            ['صافي الربح', f"{total_revenue - total_expenses:,.0f}"],
            ['مدفوعات معلقة', f"{pending_payments:,.0f}"],
        ]
        
        for row_data in financial_data:
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
                if current_row == len(financial_data) + current_row - len(financial_data):  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
            current_row += 1
        
        current_row += 2
        
        # Students Data Section
        ws[f'A{current_row}'] = "بيانات الطلاب"
        ws[f'A{current_row}'].font = sub_header_font
        current_row += 1
        
        # Students headers
        student_headers = ['#', 'اسم الطالب', 'العمر', 'الموقع', 'المجموعات', 'المدفوع', 'المتبقي', 'تاريخ التسجيل']
        for col, header in enumerate(student_headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        current_row += 1
        
        # Students data
        students = Student.query.all()
        for idx, student in enumerate(students, 1):
            groups_names = ', '.join([group.name for group in student.groups])
            student_data = [
                idx,
                student.name,
                student.age or 'غير محدد',
                student.location or 'غير محدد',
                groups_names or 'لا توجد مجموعات',
                f"{student.total_paid:,.0f}",
                f"{student.remaining_balance:,.0f}",
                student.registration_date.strftime('%Y-%m-%d') if student.registration_date else 'غير محدد'
            ]
            
            for col, value in enumerate(student_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
            current_row += 1
        
        current_row += 2
        
        # Groups Data Section
        ws[f'A{current_row}'] = "بيانات المجموعات"
        ws[f'A{current_row}'].font = sub_header_font
        current_row += 1
        
        # Groups headers
        group_headers = ['#', 'اسم المجموعة', 'المستوى', 'المدرس', 'عدد الطلاب', 'الحد الأقصى', 'السعر']
        for col, header in enumerate(group_headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        current_row += 1
        
        # Groups data
        groups = Group.query.all()
        for idx, group in enumerate(groups, 1):
            instructor_name = group.instructor_ref.name if group.instructor_ref else 'غير محدد'
            group_data = [
                idx,
                group.name,
                group.level or 'غير محدد',
                instructor_name,
                group.students.count(),
                group.max_students,
                f"{group.price:,.0f}"
            ]
            
            for col, value in enumerate(group_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
            current_row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"تقرير_شامل_{timestamp}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'حدث خطأ أثناء تصدير التقرير: {str(e)}', 'error')
        return redirect(url_for('reports'))

@app.route('/get_group_students/<int:group_id>')
def get_group_students(group_id):
    group = Group.query.get_or_404(group_id)
    student_list = []
    for student in group.students:
        student_list.append({
            'id': student.id,
            'name': student.name
        })
    return jsonify(student_list)

@app.route('/edit_student/<int:student_id>', methods=['POST'])
def edit_student(student_id):
    try:
        student = Student.query.get_or_404(student_id)
        student.name = request.form['name']
        student.phone = request.form.get('phone', '')
        student.age = int(request.form['age'])
        student.location = request.form.get('location', '')
        student.instructor_id = None  # Keep instructor optional
        registration_date = parse_date_from_input(request.form['registration_date'])
        student.discount = float(request.form.get('discount', 0))  # Get discount amount, default to 0
        
        if not registration_date:
            flash('يرجى اختيار تاريخ التسجيل من منتقي التاريخ', 'error')
            return redirect(url_for('students'))
            
        student.registration_date = registration_date
        
        # Clear existing group associations
        student.groups.clear()
        
        # Handle group selections (multiple groups allowed)
        group_ids = request.form.getlist('group_ids')  # Get list of selected group IDs
        if group_ids:
            for group_id in group_ids:
                if group_id:  # Make sure it's not empty
                    group = Group.query.get(int(group_id))
                    if group:
                        student.groups.append(group)
        
        db.session.commit()
        flash('تم تحديث بيانات الطالب بنجاح!', 'success')
        return redirect(url_for('students'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء تحديث الطالب: {str(e)}', 'error')
        return redirect(url_for('students'))

@app.route('/delete_student/<int:student_id>', methods=['POST'])
def delete_student(student_id):
    student = Student.query.get_or_404(student_id)
    
    # Delete related attendance records
    Attendance.query.filter_by(student_id=student_id).delete()
    # Delete related payment records
    Payment.query.filter_by(student_id=student_id).delete()
    
    db.session.delete(student)
    db.session.commit()
    flash('تم حذف الطالب بنجاح', 'success')
    return redirect(url_for('students'))

@app.route('/bulk_delete_students', methods=['POST'])
def bulk_delete_students():
    try:
        data = request.get_json()
        student_ids = data.get('student_ids', [])
        
        if not student_ids:
            return jsonify({'success': False, 'message': 'لم يتم تحديد أي طلاب'})
        
        # التحقق من وجود الطلاب وحذفهم
        students_deleted = 0
        for student_id in student_ids:
            student = Student.query.get(student_id)
            if student:
                # Delete related attendance records
                Attendance.query.filter_by(student_id=student_id).delete()
                # Delete related payment records
                Payment.query.filter_by(student_id=student_id).delete()
                # Delete the student
                db.session.delete(student)
                students_deleted += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'تم حذف {students_deleted} طالب بنجاح'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False, 
            'message': f'حدث خطأ أثناء الحذف: {str(e)}'
        })

@app.route('/bulk_edit_group', methods=['POST'])
def bulk_edit_group():
    try:
        student_ids = request.form.get('student_ids', '').split(',')
        student_ids = [int(id.strip()) for id in student_ids if id.strip()]
        group_id = request.form.get('group_id')
        operation = request.form.get('operation', 'add')
        
        if not student_ids:
            return jsonify({'success': False, 'message': 'لم يتم تحديد أي طلاب'})
        
        if not group_id:
            return jsonify({'success': False, 'message': 'لم يتم تحديد مجموعة'})
        
        group = Group.query.get(group_id)
        if not group:
            return jsonify({'success': False, 'message': 'المجموعة غير موجودة'})
        
        students_updated = 0
        
        for student_id in student_ids:
            student = Student.query.get(student_id)
            if not student:
                continue
                
            if operation == 'add':
                # إضافة إلى مجموعة (إذا لم يكن مضافاً بالفعل)
                if group not in student.groups:
                    student.groups.append(group)
                    students_updated += 1
                    
            elif operation == 'remove':
                # إزالة من مجموعة
                if group in student.groups:
                    student.groups.remove(group)
                    students_updated += 1
                    
            elif operation == 'replace':
                # استبدال المجموعات (إزالة الحالية وإضافة الجديدة)
                student.groups.clear()
                student.groups.append(group)
                students_updated += 1
        
        db.session.commit()
        
        operation_messages = {
            'add': f'تم إضافة {students_updated} طالب إلى مجموعة {group.name}',
            'remove': f'تم إزالة {students_updated} طالب من مجموعة {group.name}',
            'replace': f'تم تحديث مجموعات {students_updated} طالب إلى {group.name}'
        }
        
        return jsonify({
            'success': True, 
            'message': operation_messages.get(operation, 'تم التحديث بنجاح')
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False, 
            'message': f'حدث خطأ أثناء التحديث: {str(e)}'
        })

@app.route('/edit_instructor/<int:instructor_id>', methods=['POST'])
def edit_instructor(instructor_id):
    instructor = Instructor.query.get_or_404(instructor_id)
    
    instructor.name = request.form['name']
    instructor.phone = request.form['phone']
    instructor.specialization = request.form['specialization']
    
    db.session.commit()
    flash('تم تحديث بيانات المدرس بنجاح', 'success')
    return redirect(url_for('instructors'))

@app.route('/delete_instructor/<int:instructor_id>', methods=['POST'])
def delete_instructor(instructor_id):
    instructor = Instructor.query.get_or_404(instructor_id)
    
    # Check if instructor has students or groups
    if instructor.students or instructor.groups:
        flash('لا يمكن حذف المدرس لأنه مرتبط بطلاب أو مجموعات', 'error')
        return redirect(url_for('instructors'))
    
    db.session.delete(instructor)
    db.session.commit()
    flash('تم حذف المدرس بنجاح', 'success')
    return redirect(url_for('instructors'))

@app.route('/edit_group/<int:group_id>', methods=['POST'])
def edit_group(group_id):
    group = Group.query.get_or_404(group_id)
    force_save = request.form.get('force_save', 'false') == 'true'
    
    # Update basic group information
    group.name = request.form['name']
    group.level = request.form['level']
    new_instructor_id = int(request.form['instructor_id'])
    group.price = float(request.form['price'])
    group.max_students = int(request.form['max_students'])
    
    # Collect schedule data for conflict checking
    selected_days = request.form.getlist('days[]')
    schedules_to_add = []
    
    for day in selected_days:
        day_prefix = {
            'السبت': 'sat',
            'الأحد': 'sun', 
            'الاثنين': 'mon',
            'الثلاثاء': 'tue',
            'الأربعاء': 'wed',
            'الخميس': 'thu',
            'الجمعة': 'fri'
        }.get(day)
        
        if not day_prefix:
            continue
            
        hour = request.form.get(f'{day_prefix}_hour')
        minute = request.form.get(f'{day_prefix}_minute')
        period = request.form.get(f'{day_prefix}_period')
        duration = request.form.get(f'{day_prefix}_duration')
        
        if hour and minute and period and duration:
            start_time = convert_12_to_24_hour(hour, minute, period)
            
            # Calculate end time based on duration
            duration_minutes = int(duration)
            start_total_minutes = int(start_time.split(':')[0]) * 60 + int(start_time.split(':')[1])
            end_total_minutes = start_total_minutes + duration_minutes
            end_hour = (end_total_minutes // 60) % 24
            end_minute = end_total_minutes % 60
            end_time = f"{end_hour:02d}:{end_minute:02d}"
            
            schedules_to_add.append({
                'day': day,
                'start_time': start_time,
                'end_time': end_time
            })
    
    # Check for instructor schedule conflicts if not forcing save
    all_conflicts = []
    if not force_save and schedules_to_add:
        for schedule_data in schedules_to_add:
            conflicts = check_instructor_schedule_conflicts(
                schedule_data['day'], 
                schedule_data['start_time'], 
                schedule_data['end_time'], 
                new_instructor_id,
                exclude_group_id=group_id
            )
            all_conflicts.extend(conflicts)
        
        if all_conflicts:
            # Get instructor name
            instructor = Instructor.query.get(new_instructor_id)
            instructor_name = instructor.name if instructor else "غير محدد"
            
            # Return conflict information to frontend
            conflict_message = f"المدرس <strong>{instructor_name}</strong> لديه مجموعة أخرى في نفس التوقيت:<br>"
            for conflict in all_conflicts:
                start_12 = convert_24_to_12_hour(conflict['start_time'])
                end_12 = convert_24_to_12_hour(conflict['end_time'])
                conflict_message += f"• مجموعة {conflict['group_name']} - {conflict['day']}: {start_12['hour']}:{start_12['minute']} {start_12['period']} - {end_12['hour']}:{end_12['minute']} {end_12['period']}<br>"
            
            return jsonify({
                'has_conflicts': True,
                'message': conflict_message,
                'form_data': dict(request.form),
                'group_id': group_id
            })
    
    # Update instructor
    group.instructor_id = new_instructor_id
    
    # Delete existing schedules
    Schedule.query.filter_by(group_id=group_id).delete()
    
    # Add new schedules
    for schedule_data in schedules_to_add:
        schedule = Schedule(
            group_id=group.id,
            day_of_week=schedule_data['day'],
            start_time=schedule_data['start_time'],
            end_time=schedule_data['end_time']
        )
        db.session.add(schedule)
    
    db.session.commit()
    flash('تم تحديث بيانات المجموعة والجداول بنجاح', 'success')
    
    # Check if this is an AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': True, 'redirect': url_for('groups')})
    return redirect(url_for('groups'))

@app.route('/delete_group/<int:group_id>', methods=['POST'])
def delete_group(group_id):
    group = Group.query.get_or_404(group_id)
    
    # Check if group has students using the new many-to-many relationship
    student_count = group.students.count()
    if student_count > 0:
        flash('لا يمكن حذف المجموعة لأنها تحتوي على طلاب', 'error')
        return redirect(url_for('groups'))
    
    # Delete related schedules
    Schedule.query.filter_by(group_id=group_id).delete()
    # Delete related attendance records
    Attendance.query.filter_by(group_id=group_id).delete()
    
    db.session.delete(group)
    db.session.commit()
    flash('تم حذف المجموعة بنجاح', 'success')
    return redirect(url_for('groups'))

@app.route('/get_group_details/<int:group_id>')
def get_group_details(group_id):
    group = Group.query.get_or_404(group_id)
    schedules = []
    for schedule in group.schedules:
        schedules.append({
            'day': schedule.day_of_week,
            'start_time': schedule.start_time,
            'end_time': schedule.end_time
        })
    
    return jsonify({
        'id': group.id,
        'name': group.name,
        'level': group.level,
        'instructor_id': group.instructor_id,
        'max_students': group.max_students,
        'schedules': schedules
    })

@app.route('/add_expense', methods=['POST'])
def add_expense():
    description = request.form['description']
    amount = float(request.form['amount'])
    category = request.form['category']
    notes = request.form.get('notes', '')
    
    expense = Expense(
        description=description,
        amount=amount,
        category=category,
        notes=notes
    )
    
    db.session.add(expense)
    db.session.commit()
    flash('تم إضافة المصروف بنجاح', 'success')
    return redirect(url_for('payments'))

@app.route('/group_details/<int:group_id>')
@login_required
def group_details(group_id):
    group = Group.query.get_or_404(group_id)
    
    # Get all students in this group
    students = group.students.all()
    
    # Get all attendance records for this group
    attendance_records = Attendance.query.filter_by(group_id=group_id).all()
    
    # Calculate attendance statistics for the group
    total_sessions = len(set(record.date for record in attendance_records))
    total_attendances = len([record for record in attendance_records if record.status == 'حاضر'])
    total_absences = len([record for record in attendance_records if record.status == 'غائب'])
    total_late = len([record for record in attendance_records if record.status == 'متأخر'])
    
    # Calculate attendance percentage
    attendance_percentage = (total_attendances / len(attendance_records) * 100) if attendance_records else 0
    
    # Get unique dates when sessions happened
    session_dates = sorted(set(record.date for record in attendance_records), reverse=True)
    
    # Create attendance matrix for each student
    student_attendance = {}
    for student in students:
        student_records = [record for record in attendance_records if record.student_id == student.id]
        
        # Calculate student statistics
        student_present = len([r for r in student_records if r.status == 'حاضر'])
        student_absent = len([r for r in student_records if r.status == 'غائب'])
        student_late = len([r for r in student_records if r.status == 'متأخر'])
        student_percentage = (student_present / len(student_records) * 100) if student_records else 0
        
        # Create date-wise attendance - convert dates to strings for JSON serialization
        attendance_by_date = {}
        for record in student_records:
            attendance_by_date[record.date.strftime('%Y-%m-%d')] = record.status
        
        student_attendance[student.id] = {
            'student': {
                'id': student.id,
                'name': student.name,
                'phone': student.phone,
                'location': student.location
            },
            'total_present': student_present,
            'total_absent': student_absent,
            'total_late': student_late,
            'total_sessions': len(student_records),
            'percentage': round(student_percentage, 1),
            'attendance_by_date': attendance_by_date
        }
    
    # Get recent payments for this group's students
    student_ids = [s.id for s in students]
    recent_payments = Payment.query.filter(
        Payment.student_id.in_(student_ids)
    ).order_by(Payment.date.desc()).limit(10).all() if student_ids else []
    
    # Calculate financial statistics - use prices after discount
    total_expected_revenue = sum(student.total_course_price_after_discount for student in students)
    total_received_revenue = sum(student.total_paid for student in students)
    pending_revenue = total_expected_revenue - total_received_revenue
    
    return render_template('group_details.html',
                         group=group,
                         students=students,
                         session_dates=[date.strftime('%Y-%m-%d') for date in session_dates],
                         student_attendance=student_attendance,
                         total_sessions=total_sessions,
                         total_attendances=total_attendances,
                         total_absences=total_absences,
                         total_late=total_late,
                         attendance_percentage=round(attendance_percentage, 1),
                         recent_payments=recent_payments,
                         total_expected_revenue=total_expected_revenue,
                         total_received_revenue=total_received_revenue,
                         pending_revenue=pending_revenue)

@app.route('/add_sample_attendance')
@admin_required
def add_sample_attendance():
    """Add sample attendance data for testing - Admin only"""
    from datetime import date, timedelta
    
    # Get all groups and their students
    groups = Group.query.all()
    
    # Generate attendance for the last 30 days
    start_date = date.today() - timedelta(days=30)
    
    for group in groups:
        students = group.students.all()
        if not students:
            continue
            
        # Generate attendance for each day in the last 30 days
        for i in range(30):
            current_date = start_date + timedelta(days=i)
            
            # Skip weekends (Friday and Saturday in Middle East)
            if current_date.weekday() in [4, 5]:  # Friday and Saturday
                continue
                
            for student in students:
                # Check if attendance already exists
                existing = Attendance.query.filter_by(
                    student_id=student.id,
                    date=current_date,
                    group_id=group.id
                ).first()
                
                if not existing:
                    # Generate random attendance status
                    # 70% present, 20% absent, 10% late
                    rand = random.random()
                    if rand < 0.7:
                        status = 'حاضر'
                    elif rand < 0.9:
                        status = 'غائب'
                    else:
                        status = 'متأخر'
                    
                    attendance = Attendance(
                        student_id=student.id,
                        date=current_date,
                        status=status,
                        group_id=group.id
                    )
                    db.session.add(attendance)
    
    db.session.commit()
    flash('تم إضافة بيانات الحضور التجريبية بنجاح!', 'success')
    return redirect(url_for('groups'))

def convert_12_to_24_hour(hour, minute, period):
    """Convert 12-hour format to 24-hour format"""
    hour = int(hour)
    minute = int(minute)
    
    if period == 'AM':
        if hour == 12:
            hour = 0
    else:  # PM
        if hour != 12:
            hour += 12
    
    return f"{hour:02d}:{minute:02d}"

def convert_24_to_12_hour(time_24):
    """Convert 24-hour format to 12-hour format"""
    if not time_24 or ':' not in time_24:
        return {'hour': '12', 'minute': '00', 'period': 'AM'}
    
    hour, minute = time_24.split(':')
    hour = int(hour)
    period = 'AM'
    
    if hour == 0:
        hour = 12
    elif hour == 12:
        period = 'PM'
    elif hour > 12:
        hour = hour - 12
        period = 'PM'
    
    return {'hour': str(hour), 'minute': minute, 'period': period}

# Add the function to Jinja2 template context
@app.context_processor
def utility_processor():
    def get_new_instructor_notes_count():
        """Get count of new instructor notes for admin notification"""
        current_user = get_current_user()
        if current_user and current_user.role == 'admin':
            return InstructorNote.query.filter_by(status='جديد').count()
        return 0
    
    return dict(
        convert_24_to_12_hour=convert_24_to_12_hour,
        get_arabic_day_name=get_arabic_day_name,
        format_arabic_date=format_arabic_date,
        format_time_12hour=format_time_12hour,
        format_date_for_input=format_date_for_input,
        get_new_instructor_notes_count=get_new_instructor_notes_count
    )

def init_db():
    """Initialize database and create default admin"""
    with app.app_context():
        db.create_all()
        create_default_admin()

@app.route('/debug')
@login_required
def debug_prices():
    """Debug page to test price calculation"""
    return render_template('debug.html')

@app.route('/tasks')
@login_required
def tasks():
    """Display tasks and notes management page"""
    filter_status = request.args.get('status', 'all')
    filter_priority = request.args.get('priority', 'all')
    filter_category = request.args.get('category', 'all')
    
    # Build query based on filters for tasks
    query = Task.query
    
    if filter_status != 'all':
        query = query.filter_by(status=filter_status)
    
    if filter_priority != 'all':
        query = query.filter_by(priority=filter_priority)
    
    # Order by priority and creation date
    priority_order = {'عالي': 3, 'متوسط': 2, 'منخفض': 1}
    tasks = query.all()
    tasks.sort(key=lambda x: (priority_order.get(x.priority, 0), x.created_at), reverse=True)
    
    # Get notes and filter them
    notes_query = Note.query
    
    if filter_category != 'all':
        notes_query = notes_query.filter_by(category=filter_category)
    
    # Order notes by pinned status and creation date
    notes = notes_query.order_by(Note.is_pinned.desc(), Note.updated_at.desc()).all()
    
    # Get instructor notes (for admins only)
    instructor_notes = []
    current_user = get_current_user()
    if current_user.role == 'admin':
        instructor_notes_query = InstructorNote.query
        
        if filter_status != 'all':
            instructor_notes_query = instructor_notes_query.filter_by(status=filter_status)
        
        if filter_priority != 'all':
            instructor_notes_query = instructor_notes_query.filter_by(priority=filter_priority)
        
        instructor_notes = instructor_notes_query.order_by(InstructorNote.created_at.desc()).all()
    
    # Get statistics
    total_tasks = Task.query.count()
    completed_tasks = Task.query.filter_by(status='مكتمل').count()
    pending_tasks = Task.query.filter(Task.status.in_(['قيد التنفيذ'])).count()
    overdue_tasks = len([t for t in Task.query.all() if t.is_overdue])
    
    # Notes statistics
    total_notes = Note.query.count()
    pinned_notes = Note.query.filter_by(is_pinned=True).count()
    
    # Instructor notes statistics (for admins)
    total_instructor_notes = 0
    new_instructor_notes = 0
    if current_user.role == 'admin':
        total_instructor_notes = InstructorNote.query.count()
        new_instructor_notes = InstructorNote.query.filter_by(status='جديد').count()
    
    users = User.query.all()
    
    return render_template('tasks.html',
                         tasks=tasks,
                         notes=notes,
                         instructor_notes=instructor_notes,
                         users=users,
                         current_user=current_user,
                         total_tasks=total_tasks,
                         completed_tasks=completed_tasks,
                         pending_tasks=pending_tasks,
                         overdue_tasks=overdue_tasks,
                         total_notes=total_notes,
                         pinned_notes=pinned_notes,
                         total_instructor_notes=total_instructor_notes,
                         new_instructor_notes=new_instructor_notes,
                         filter_status=filter_status,
                         filter_priority=filter_priority,
                         filter_category=filter_category)

@app.route('/add_task', methods=['POST'])
@login_required
def add_task():
    """Add a new task"""
    try:
        title = request.form['title']
        description = request.form.get('description', '')
        priority = request.form['priority']
        due_date_str = request.form.get('due_date')
        assigned_to = request.form.get('assigned_to')
        
        # Parse due date
        due_date = None
        if due_date_str:
            due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
        
        # Convert assigned_to to int if provided
        assigned_to_id = None
        if assigned_to and assigned_to != '':
            assigned_to_id = int(assigned_to)
        
        current_user = get_current_user()
        
        task = Task(
            title=title,
            description=description,
            priority=priority,
            due_date=due_date,
            created_by=current_user.id,
            assigned_to=assigned_to_id
        )
        
        db.session.add(task)
        db.session.commit()
        
        flash('تم إضافة المهمة بنجاح!', 'success')
        return redirect(url_for('tasks'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء إضافة المهمة: {str(e)}', 'error')
        return redirect(url_for('tasks'))

@app.route('/update_task_status/<int:task_id>', methods=['POST'])
@login_required
def update_task_status(task_id):
    """Update task status"""
    try:
        task = Task.query.get_or_404(task_id)
        new_status = request.form['status']
        
        task.status = new_status
        
        # If marking as completed, set completion time
        if new_status == 'مكتمل':
            task.completed_at = datetime.utcnow()
        else:
            task.completed_at = None
        
        db.session.commit()
        
        flash('تم تحديث حالة المهمة بنجاح!', 'success')
        return redirect(url_for('tasks'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء تحديث المهمة: {str(e)}', 'error')
        return redirect(url_for('tasks'))

@app.route('/edit_task/<int:task_id>', methods=['POST'])
@login_required
def edit_task(task_id):
    """Edit an existing task"""
    try:
        task = Task.query.get_or_404(task_id)
        
        task.title = request.form['title']
        task.description = request.form.get('description', '')
        task.priority = request.form['priority']
        
        # Parse due date
        due_date_str = request.form.get('due_date')
        if due_date_str:
            task.due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
        else:
            task.due_date = None
        
        # Update assigned user
        assigned_to = request.form.get('assigned_to')
        if assigned_to and assigned_to != '':
            task.assigned_to = int(assigned_to)
        else:
            task.assigned_to = None
        
        db.session.commit()
        
        flash('تم تحديث المهمة بنجاح!', 'success')
        return redirect(url_for('tasks'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء تحديث المهمة: {str(e)}', 'error')
        return redirect(url_for('tasks'))

@app.route('/delete_task/<int:task_id>', methods=['POST'])
@login_required
def delete_task(task_id):
    """Delete a task"""
    try:
        task = Task.query.get_or_404(task_id)
        db.session.delete(task)
        db.session.commit()
        
        flash('تم حذف المهمة بنجاح!', 'success')
        return redirect(url_for('tasks'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف المهمة: {str(e)}', 'error')
        return redirect(url_for('tasks'))

@app.route('/add_note', methods=['POST'])
@login_required
def add_note():
    """Add a new note"""
    try:
        title = request.form['title']
        content = request.form['content']
        category = request.form['category']
        color = request.form['color']
        is_pinned = 'is_pinned' in request.form
        
        current_user = get_current_user()
        
        note = Note(
            title=title,
            content=content,
            category=category,
            color=color,
            is_pinned=is_pinned,
            created_by=current_user.id
        )
        
        db.session.add(note)
        db.session.commit()
        
        flash('تم إضافة الملاحظة بنجاح!', 'success')
        return redirect(url_for('tasks') + '#notes')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء إضافة الملاحظة: {str(e)}', 'error')
        return redirect(url_for('tasks') + '#notes')

@app.route('/edit_note/<int:note_id>', methods=['POST'])
@login_required
def edit_note(note_id):
    """Edit an existing note"""
    try:
        note = Note.query.get_or_404(note_id)
        
        note.title = request.form['title']
        note.content = request.form['content']
        note.category = request.form['category']
        note.color = request.form['color']
        note.is_pinned = 'is_pinned' in request.form
        note.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        flash('تم تحديث الملاحظة بنجاح!', 'success')
        return redirect(url_for('tasks') + '#notes')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء تحديث الملاحظة: {str(e)}', 'error')
        return redirect(url_for('tasks') + '#notes')

@app.route('/delete_note/<int:note_id>', methods=['POST'])
@login_required
def delete_note(note_id):
    """Delete a note"""
    try:
        note = Note.query.get_or_404(note_id)
        db.session.delete(note)
        db.session.commit()
        
        flash('تم حذف الملاحظة بنجاح!', 'success')
        return redirect(url_for('tasks') + '#notes')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف الملاحظة: {str(e)}', 'error')
        return redirect(url_for('tasks') + '#notes')

@app.route('/toggle_pin_note/<int:note_id>', methods=['POST'])
@login_required
def toggle_pin_note(note_id):
    """Toggle pin status of a note"""
    try:
        note = Note.query.get_or_404(note_id)
        note.is_pinned = not note.is_pinned
        note.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        flash('تم تحديث تثبيت الملاحظة!', 'success')
        return redirect(url_for('tasks') + '#notes')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء تحديث الملاحظة: {str(e)}', 'error')
        return redirect(url_for('tasks') + '#notes')

# Health check endpoints for monitoring
@app.route('/health')
def health_check():
    """Basic health check endpoint"""
    try:
        # Test database connection
        db.session.execute('SELECT 1')
        db_status = 'healthy'
    except Exception:
        db_status = 'unhealthy'
    
    return jsonify({
        'status': 'healthy' if db_status == 'healthy' else 'unhealthy',
        'database': db_status,
        'timestamp': datetime.utcnow().isoformat(),
        'version': '1.0.0'
    }), 200 if db_status == 'healthy' else 503

@app.route('/ping')
def ping():
    """Simple ping endpoint"""
    return 'pong', 200

@app.route('/status')
def status():
    """Detailed status information"""
    try:
        # Check database
        db.session.execute('SELECT 1')
        db_status = 'connected'
        
        # Count records
        users_count = User.query.count()
        students_count = Student.query.count()
        
    except Exception as e:
        db_status = f'error: {str(e)}'
        users_count = -1
        students_count = -1
    
    return jsonify({
        'app_name': 'Tafra Student Management System',
        'version': '1.0.0',
        'status': 'running',
        'environment': os.environ.get('FLASK_ENV', 'development'),
        'database': {
            'status': db_status,
            'users_count': users_count,
            'students_count': students_count
        },
        'timestamp': datetime.utcnow().isoformat()
    }), 200

def get_instructor_groups(user):
    """Get groups assigned to a specific instructor user"""
    if user.role == 'admin':
        return Group.query.all()
    elif user.role == 'instructor' and user.linked_instructor:
        return user.linked_instructor.groups
    return []

def get_instructor_students(user):
    """Get students assigned to a specific instructor user"""
    if user.role == 'admin':
        return Student.query.all()
    elif user.role == 'instructor' and user.linked_instructor:
        # Get all students in instructor's groups
        instructor_groups = user.linked_instructor.groups
        students = set()
        for group in instructor_groups:
            students.update(group.students)
        return list(students)
    return []

@app.route('/instructor_attendance')
@instructor_required
def instructor_attendance():
    current_user = get_current_user()
    instructor_groups = get_instructor_groups(current_user)
    
    # Get attendance for instructor's groups only
    today = datetime.now().date()
    attendance_records = []
    
    for group in instructor_groups:
        group_attendance = Attendance.query.filter_by(
            group_id=group.id,
            date=today
        ).all()
        attendance_records.extend(group_attendance)
    
    return render_template('instructor_attendance.html',
                         groups=instructor_groups,
                         attendance_records=attendance_records,
                         today=today)

@app.route('/instructor_mark_attendance', methods=['POST'])
@instructor_required
def instructor_mark_attendance():
    current_user = get_current_user()
    instructor_groups = get_instructor_groups(current_user)
    group_ids = [g.id for g in instructor_groups]
    
    group_id = int(request.form['group_id'])
    
    # Verify instructor has access to this group
    if group_id not in group_ids:
        flash('ليس لديك صلاحية لأخذ حضور هذه المجموعة', 'error')
        return redirect(url_for('instructor_attendance'))
    
    date_str = request.form['date']
    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    
    group = Group.query.get(group_id)
    students = group.students
    
    for student in students:
        student_id = str(student.id)
        status = request.form.get(f'attendance_{student_id}')
        
        if status:
            # Check if attendance already exists
            existing = Attendance.query.filter_by(
                student_id=student.id,
                date=date_obj,
                group_id=group_id
            ).first()
            
            if existing:
                existing.status = status
            else:
                attendance = Attendance(
                    student_id=student.id,
                    date=date_obj,
                    status=status,
                    group_id=group_id
                )
                db.session.add(attendance)
    
    try:
        db.session.commit()
        flash('تم حفظ الحضور بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash('حدث خطأ أثناء حفظ الحضور', 'error')
    
    return redirect(url_for('instructor_attendance'))

@app.route('/instructor_notes')
@instructor_required
def instructor_notes():
    current_user = get_current_user()
    instructor_groups = get_instructor_groups(current_user)
    instructor_students = get_instructor_students(current_user)
    
    # Get instructor's notes
    notes = InstructorNote.query.filter_by(created_by=current_user.id)\
                               .order_by(InstructorNote.created_at.desc()).all()
    
    return render_template('instructor_notes.html',
                         notes=notes,
                         instructor_groups=instructor_groups,
                         instructor_students=instructor_students)

@app.route('/add_instructor_note', methods=['POST'])
@instructor_required
def add_instructor_note():
    current_user = get_current_user()
    
    title = request.form['title']
    content = request.form['content']
    priority = request.form['priority']
    student_id = request.form.get('student_id') if request.form.get('student_id') else None
    group_id = request.form.get('group_id') if request.form.get('group_id') else None
    
    # Verify instructor has access to selected student/group
    if student_id:
        instructor_students = get_instructor_students(current_user)
        student_ids = [s.id for s in instructor_students]
        if int(student_id) not in student_ids:
            flash('ليس لديك صلاحية لإضافة ملاحظة لهذا الطالب', 'error')
            return redirect(url_for('instructor_notes'))
    
    if group_id:
        instructor_groups = get_instructor_groups(current_user)
        group_ids = [g.id for g in instructor_groups]
        if int(group_id) not in group_ids:
            flash('ليس لديك صلاحية لإضافة ملاحظة لهذه المجموعة', 'error')
            return redirect(url_for('instructor_notes'))
    
    note = InstructorNote(
        title=title,
        content=content,
        priority=priority,
        student_id=int(student_id) if student_id else None,
        group_id=int(group_id) if group_id else None,
        created_by=current_user.id
    )
    
    try:
        db.session.add(note)
        db.session.commit()
        flash('تم إرسال الملاحظة للإدارة بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash('حدث خطأ أثناء إضافة الملاحظة', 'error')
    
    return redirect(url_for('instructor_notes'))

@app.route('/instructor_todos')
@instructor_required
def instructor_todos():
    """Display instructor's personal todo list"""
    current_user = get_current_user()
    instructor_groups = get_instructor_groups(current_user)
    instructor_students = get_instructor_students(current_user)
    
    # Get filter parameters
    filter_status = request.args.get('status', 'all')
    filter_priority = request.args.get('priority', 'all')
    filter_category = request.args.get('category', 'all')
    
    # Build query based on filters
    query = InstructorTodo.query.filter_by(created_by=current_user.id)
    
    if filter_status != 'all':
        query = query.filter_by(status=filter_status)
    
    if filter_priority != 'all':
        query = query.filter_by(priority=filter_priority)
    
    if filter_category != 'all':
        query = query.filter_by(category=filter_category)
    
    # Order by priority and creation date
    priority_order = {'عالي': 3, 'متوسط': 2, 'منخفض': 1}
    todos = query.all()
    todos.sort(key=lambda x: (priority_order.get(x.priority, 0), x.created_at), reverse=True)
    
    # Get statistics
    total_todos = InstructorTodo.query.filter_by(created_by=current_user.id).count()
    open_todos = InstructorTodo.query.filter_by(created_by=current_user.id, status='مفتوح').count()
    completed_todos = InstructorTodo.query.filter_by(created_by=current_user.id, status='مكتمل').count()
    overdue_todos = len([t for t in InstructorTodo.query.filter_by(created_by=current_user.id).all() if t.is_overdue])
    
    return render_template('instructor_todos.html',
                         todos=todos,
                         instructor_groups=instructor_groups,
                         instructor_students=instructor_students,
                         total_todos=total_todos,
                         open_todos=open_todos,
                         completed_todos=completed_todos,
                         overdue_todos=overdue_todos,
                         filter_status=filter_status,
                         filter_priority=filter_priority,
                         filter_category=filter_category)

@app.route('/add_instructor_todo', methods=['POST'])
@instructor_required
def add_instructor_todo():
    """Add a new todo item for instructor"""
    try:
        current_user = get_current_user()
        
        title = request.form['title']
        description = request.form.get('description', '')
        category = request.form['category']
        priority = request.form['priority']
        due_date_str = request.form.get('due_date')
        group_id = request.form.get('group_id')
        student_id = request.form.get('student_id')
        
        # Parse due date
        due_date = None
        if due_date_str:
            due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
        
        # Verify instructor has access to selected group/student
        if group_id:
            instructor_groups = get_instructor_groups(current_user)
            group_ids = [g.id for g in instructor_groups]
            if int(group_id) not in group_ids:
                flash('ليس لديك صلاحية لربط المهمة بهذه المجموعة', 'error')
                return redirect(url_for('instructor_todos'))
        
        if student_id:
            instructor_students = get_instructor_students(current_user)
            student_ids = [s.id for s in instructor_students]
            if int(student_id) not in student_ids:
                flash('ليس لديك صلاحية لربط المهمة بهذا الطالب', 'error')
                return redirect(url_for('instructor_todos'))
        
        todo = InstructorTodo(
            title=title,
            description=description,
            category=category,
            priority=priority,
            due_date=due_date,
            group_id=int(group_id) if group_id else None,
            student_id=int(student_id) if student_id else None,
            created_by=current_user.id
        )
        
        db.session.add(todo)
        db.session.commit()
        
        flash('تم إضافة المهمة بنجاح!', 'success')
        return redirect(url_for('instructor_todos'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء إضافة المهمة: {str(e)}', 'error')
        return redirect(url_for('instructor_todos'))

@app.route('/update_instructor_todo_status/<int:todo_id>', methods=['POST'])
@instructor_required
def update_instructor_todo_status(todo_id):
    """Update todo status"""
    try:
        current_user = get_current_user()
        todo = InstructorTodo.query.filter_by(id=todo_id, created_by=current_user.id).first_or_404()
        
        new_status = request.form['status']
        todo.status = new_status
        
        # If marking as completed, set completion time
        if new_status == 'مكتمل':
            todo.completed_at = datetime.utcnow()
        else:
            todo.completed_at = None
        
        todo.updated_at = datetime.utcnow()
        db.session.commit()
        
        flash('تم تحديث حالة المهمة بنجاح!', 'success')
        return redirect(url_for('instructor_todos'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء تحديث المهمة: {str(e)}', 'error')
        return redirect(url_for('instructor_todos'))

@app.route('/edit_instructor_todo/<int:todo_id>', methods=['POST'])
@instructor_required
def edit_instructor_todo(todo_id):
    """Edit an existing todo"""
    try:
        current_user = get_current_user()
        todo = InstructorTodo.query.filter_by(id=todo_id, created_by=current_user.id).first_or_404()
        
        todo.title = request.form['title']
        todo.description = request.form.get('description', '')
        todo.category = request.form['category']
        todo.priority = request.form['priority']
        
        # Parse due date
        due_date_str = request.form.get('due_date')
        if due_date_str:
            todo.due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
        else:
            todo.due_date = None
        
        # Update group and student
        group_id = request.form.get('group_id')
        student_id = request.form.get('student_id')
        
        # Verify access
        if group_id:
            instructor_groups = get_instructor_groups(current_user)
            group_ids = [g.id for g in instructor_groups]
            if int(group_id) not in group_ids:
                flash('ليس لديك صلاحية لربط المهمة بهذه المجموعة', 'error')
                return redirect(url_for('instructor_todos'))
            todo.group_id = int(group_id)
        else:
            todo.group_id = None
        
        if student_id:
            instructor_students = get_instructor_students(current_user)
            student_ids = [s.id for s in instructor_students]
            if int(student_id) not in student_ids:
                flash('ليس لديك صلاحية لربط المهمة بهذا الطالب', 'error')
                return redirect(url_for('instructor_todos'))
            todo.student_id = int(student_id)
        else:
            todo.student_id = None
        
        todo.updated_at = datetime.utcnow()
        db.session.commit()
        
        flash('تم تحديث المهمة بنجاح!', 'success')
        return redirect(url_for('instructor_todos'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء تحديث المهمة: {str(e)}', 'error')
        return redirect(url_for('instructor_todos'))

@app.route('/delete_instructor_todo/<int:todo_id>', methods=['POST'])
@instructor_required
def delete_instructor_todo(todo_id):
    """Delete a todo"""
    try:
        current_user = get_current_user()
        todo = InstructorTodo.query.filter_by(id=todo_id, created_by=current_user.id).first_or_404()
        
        db.session.delete(todo)
        db.session.commit()
        
        flash('تم حذف المهمة بنجاح!', 'success')
        return redirect(url_for('instructor_todos'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف المهمة: {str(e)}', 'error')
        return redirect(url_for('instructor_todos'))

@app.route('/get_instructor_todo/<int:todo_id>')
@instructor_required
def get_instructor_todo(todo_id):
    """Get todo details for editing"""
    try:
        current_user = get_current_user()
        todo = InstructorTodo.query.filter_by(id=todo_id, created_by=current_user.id).first_or_404()
        
        return jsonify({
            'id': todo.id,
            'title': todo.title,
            'description': todo.description,
            'category': todo.category,
            'priority': todo.priority,
            'due_date': todo.due_date.strftime('%Y-%m-%d') if todo.due_date else '',
            'group_id': todo.group_id,
            'student_id': todo.student_id,
            'status': todo.status
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 404

@app.route('/export_full_backup')
@admin_required
def export_full_backup():
    """Export complete system backup with all data"""
    try:
        # Create workbook
        wb = Workbook()
        
        # Set RTL direction for Arabic support
        ws_overview = wb.active
        ws_overview.title = "نظرة عامة"
        ws_overview.sheet_view.rightToLeft = True
        
        # Define styles
        header_font = Font(size=14, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        sub_header_font = Font(size=12, bold=True, color="2F5F8F")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        current_row = 1
        
        # Title and backup info
        ws_overview.merge_cells(f'A{current_row}:H{current_row}')
        title_cell = ws_overview[f'A{current_row}']
        title_cell.value = f"نسخة احتياطية شاملة - نظام تفرا لإدارة الطلاب - {format_arabic_date(datetime.now())}"
        title_cell.font = Font(size=16, bold=True, color="2F5F8F")
        title_cell.alignment = center_alignment
        current_row += 2
        
        # System overview
        ws_overview[f'A{current_row}'] = "معلومات النظام"
        ws_overview[f'A{current_row}'].font = sub_header_font
        current_row += 1
        
        overview_data = [
            ['البيان', 'القيمة'],
            ['تاريخ النسخة الاحتياطية', format_arabic_date(datetime.now())],
            ['وقت النسخة الاحتياطية', datetime.now().strftime('%H:%M:%S')],
            ['إجمالي الطلاب', Student.query.count()],
            ['إجمالي المدرسين', Instructor.query.count()],
            ['إجمالي المجموعات', Group.query.count()],
            ['إجمالي المستخدمين', User.query.count()],
            ['إجمالي المدفوعات', Payment.query.count()],
            ['إجمالي المصروفات', Expense.query.count()],
            ['إجمالي سجلات الحضور', Attendance.query.count()],
            ['إجمالي المهام', Task.query.count()],
            ['إجمالي الملاحظات', Note.query.count()],
            ['إجمالي ملاحظات المدرسين', InstructorNote.query.count()],
            ['إجمالي مهام المدرسين', InstructorTodo.query.count()],
        ]
        
        for row_data in overview_data:
            for col, value in enumerate(row_data, 1):
                cell = ws_overview.cell(row=current_row, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
                if current_row == len(overview_data) + current_row - len(overview_data):  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
            current_row += 1
        
        current_row += 2
        
        # Users data sheet
        ws_users = wb.create_sheet(title="المستخدمين")
        ws_users.sheet_view.rightToLeft = True
        
        user_headers = ['#', 'اسم المستخدم', 'الاسم الكامل', 'الدور', 'مخفي', 'تاريخ الإنشاء', 'آخر دخول', 'آخر نشاط', 'نشط الآن']
        for col, header in enumerate(user_headers, 1):
            cell = ws_users.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        users = User.query.all()
        for idx, user in enumerate(users, 1):
            user_data = [
                idx,
                user.username,
                user.full_name,
                user.role,
                'نعم' if user.is_hidden else 'لا',
                user.created_at.strftime('%Y-%m-%d %H:%M') if user.created_at else 'غير محدد',
                user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'لم يسجل دخول',
                user.last_activity.strftime('%Y-%m-%d %H:%M') if user.last_activity else 'غير محدد',
                'نعم' if user.is_active_now() else 'لا'
            ]
            
            for col, value in enumerate(user_data, 1):
                cell = ws_users.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Students data sheet
        ws_students = wb.create_sheet(title="الطلاب")
        ws_students.sheet_view.rightToLeft = True
        
        student_headers = ['#', 'اسم الطالب', 'الهاتف', 'العمر', 'الموقع', 'المدرس', 'المجموعات', 'إجمالي السعر', 'الخصم', 'السعر بعد الخصم', 'المدفوع', 'المتبقي', 'تاريخ التسجيل']
        for col, header in enumerate(student_headers, 1):
            cell = ws_students.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        students = Student.query.all()
        for idx, student in enumerate(students, 1):
            instructor_name = student.instructor_ref.name if student.instructor_ref else 'غير محدد'
            groups_names = ', '.join([group.name for group in student.groups])
            
            student_data = [
                idx,
                student.name,
                student.phone or 'غير محدد',
                student.age or 'غير محدد',
                student.location or 'غير محدد',
                instructor_name,
                groups_names or 'لا توجد مجموعات',
                f"{student.total_course_price:,.0f}",
                f"{student.discount:,.0f}",
                f"{student.total_course_price_after_discount:,.0f}",
                f"{student.total_paid:,.0f}",
                f"{student.remaining_balance:,.0f}",
                student.registration_date.strftime('%Y-%m-%d') if student.registration_date else 'غير محدد'
            ]
            
            for col, value in enumerate(student_data, 1):
                cell = ws_students.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Instructors data sheet
        ws_instructors = wb.create_sheet(title="المدرسين")
        ws_instructors.sheet_view.rightToLeft = True
        
        instructor_headers = ['#', 'اسم المدرس', 'الهاتف', 'التخصص', 'عدد الطلاب', 'عدد المجموعات', 'مرتبط بمستخدم']
        for col, header in enumerate(instructor_headers, 1):
            cell = ws_instructors.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        instructors = Instructor.query.all()
        for idx, instructor in enumerate(instructors, 1):
            linked_user = 'نعم' if instructor.user_account else 'لا'
            
            instructor_data = [
                idx,
                instructor.name,
                instructor.phone or 'غير محدد',
                instructor.specialization or 'غير محدد',
                len(instructor.students),
                len(instructor.groups),
                linked_user
            ]
            
            for col, value in enumerate(instructor_data, 1):
                cell = ws_instructors.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Groups data sheet
        ws_groups = wb.create_sheet(title="المجموعات")
        ws_groups.sheet_view.rightToLeft = True
        
        group_headers = ['#', 'اسم المجموعة', 'المستوى', 'المدرس', 'عدد الطلاب', 'الحد الأقصى', 'السعر', 'أيام الدروس', 'أوقات الدروس']
        for col, header in enumerate(group_headers, 1):
            cell = ws_groups.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        groups = Group.query.all()
        for idx, group in enumerate(groups, 1):
            instructor_name = group.instructor_ref.name if group.instructor_ref else 'غير محدد'
            
            # Get schedule details
            schedules = []
            for schedule in group.schedules:
                start_12 = convert_24_to_12_hour(schedule.start_time)
                end_12 = convert_24_to_12_hour(schedule.end_time)
                schedules.append(f"{schedule.day_of_week}: {start_12['hour']}:{start_12['minute']} {start_12['period']} - {end_12['hour']}:{end_12['minute']} {end_12['period']}")
            
            days = ', '.join([s.day_of_week for s in group.schedules])
            times = ' | '.join(schedules)
            
            group_data = [
                idx,
                group.name,
                group.level or 'غير محدد',
                instructor_name,
                group.students.count(),
                group.max_students,
                f"{group.price:,.0f}",
                days or 'غير محدد',
                times or 'غير محدد'
            ]
            
            for col, value in enumerate(group_data, 1):
                cell = ws_groups.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Schedules data sheet
        ws_schedules = wb.create_sheet(title="الجداول الزمنية")
        ws_schedules.sheet_view.rightToLeft = True
        
        schedule_headers = ['#', 'المجموعة', 'المدرس', 'اليوم', 'وقت البداية', 'وقت النهاية', 'المدة']
        for col, header in enumerate(schedule_headers, 1):
            cell = ws_schedules.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        schedules = Schedule.query.all()
        for idx, schedule in enumerate(schedules, 1):
            group = Group.query.get(schedule.group_id)
            instructor_name = group.instructor_ref.name if group and group.instructor_ref else 'غير محدد'
            group_name = group.name if group else 'مجموعة محذوفة'
            
            # Calculate duration
            try:
                start_time = datetime.strptime(schedule.start_time, '%H:%M').time()
                end_time = datetime.strptime(schedule.end_time, '%H:%M').time()
                start_datetime = datetime.combine(datetime.today(), start_time)
                end_datetime = datetime.combine(datetime.today(), end_time)
                duration = end_datetime - start_datetime
                duration_str = f"{duration.seconds // 3600}:{(duration.seconds % 3600) // 60:02d}"
            except:
                duration_str = 'غير محدد'
            
            schedule_data = [
                idx,
                group_name,
                instructor_name,
                schedule.day_of_week,
                schedule.start_time,
                schedule.end_time,
                duration_str
            ]
            
            for col, value in enumerate(schedule_data, 1):
                cell = ws_schedules.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Payments data sheet
        ws_payments = wb.create_sheet(title="المدفوعات")
        ws_payments.sheet_view.rightToLeft = True
        
        payment_headers = ['#', 'اسم الطالب', 'المبلغ', 'الشهر', 'التاريخ', 'ملاحظات']
        for col, header in enumerate(payment_headers, 1):
            cell = ws_payments.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        payments = Payment.query.order_by(Payment.date.desc()).all()
        for idx, payment in enumerate(payments, 1):
            student = Student.query.get(payment.student_id)
            student_name = student.name if student else 'طالب محذوف'
            
            payment_data = [
                idx,
                student_name,
                f"{payment.amount:,.0f}",
                payment.month or 'غير محدد',
                payment.date.strftime('%Y-%m-%d %H:%M') if payment.date else 'غير محدد',
                payment.notes or 'لا توجد ملاحظات'
            ]
            
            for col, value in enumerate(payment_data, 1):
                cell = ws_payments.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Expenses data sheet
        ws_expenses = wb.create_sheet(title="المصروفات")
        ws_expenses.sheet_view.rightToLeft = True
        
        expense_headers = ['#', 'الوصف', 'المبلغ', 'الفئة', 'التاريخ', 'ملاحظات']
        for col, header in enumerate(expense_headers, 1):
            cell = ws_expenses.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        expenses = Expense.query.order_by(Expense.date.desc()).all()
        for idx, expense in enumerate(expenses, 1):
            expense_data = [
                idx,
                expense.description,
                f"{expense.amount:,.0f}",
                expense.category or 'غير محدد',
                expense.date.strftime('%Y-%m-%d %H:%M') if expense.date else 'غير محدد',
                expense.notes or 'لا توجد ملاحظات'
            ]
            
            for col, value in enumerate(expense_data, 1):
                cell = ws_expenses.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Attendance data sheet (last 30 days)
        ws_attendance = wb.create_sheet(title="الحضور")
        ws_attendance.sheet_view.rightToLeft = True
        
        attendance_headers = ['#', 'اسم الطالب', 'المجموعة', 'التاريخ', 'الحالة']
        for col, header in enumerate(attendance_headers, 1):
            cell = ws_attendance.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        # Get attendance for last 30 days
        thirty_days_ago = datetime.now().date() - timedelta(days=30)
        attendance_records = Attendance.query.filter(Attendance.date >= thirty_days_ago).order_by(Attendance.date.desc()).all()
        
        for idx, record in enumerate(attendance_records, 1):
            student = Student.query.get(record.student_id)
            group = Group.query.get(record.group_id)
            student_name = student.name if student else 'طالب محذوف'
            group_name = group.name if group else 'مجموعة محذوفة'
            
            attendance_data = [
                idx,
                student_name,
                group_name,
                record.date.strftime('%Y-%m-%d') if record.date else 'غير محدد',
                record.status
            ]
            
            for col, value in enumerate(attendance_data, 1):
                cell = ws_attendance.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Tasks data sheet
        ws_tasks = wb.create_sheet(title="المهام")
        ws_tasks.sheet_view.rightToLeft = True
        
        task_headers = ['#', 'العنوان', 'الوصف', 'الأولوية', 'الحالة', 'تاريخ الاستحقاق', 'منشئ المهمة', 'المُكلف', 'تاريخ الإنشاء', 'تاريخ الإكمال']
        for col, header in enumerate(task_headers, 1):
            cell = ws_tasks.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        tasks = Task.query.order_by(Task.created_at.desc()).all()
        for idx, task in enumerate(tasks, 1):
            creator = User.query.get(task.created_by)
            assignee = User.query.get(task.assigned_to) if task.assigned_to else None
            
            task_data = [
                idx,
                task.title,
                task.description or 'لا يوجد وصف',
                task.priority,
                task.status,
                task.due_date.strftime('%Y-%m-%d') if task.due_date else 'غير محدد',
                creator.full_name if creator else 'مستخدم محذوف',
                assignee.full_name if assignee else 'غير مُكلف',
                task.created_at.strftime('%Y-%m-%d %H:%M') if task.created_at else 'غير محدد',
                task.completed_at.strftime('%Y-%m-%d %H:%M') if task.completed_at else 'غير مكتمل'
            ]
            
            for col, value in enumerate(task_data, 1):
                cell = ws_tasks.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Notes data sheet
        ws_notes = wb.create_sheet(title="الملاحظات")
        ws_notes.sheet_view.rightToLeft = True
        
        note_headers = ['#', 'العنوان', 'المحتوى', 'الفئة', 'اللون', 'مثبت', 'منشئ الملاحظة', 'تاريخ الإنشاء', 'تاريخ التحديث']
        for col, header in enumerate(note_headers, 1):
            cell = ws_notes.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        notes = Note.query.order_by(Note.updated_at.desc()).all()
        for idx, note in enumerate(notes, 1):
            creator = User.query.get(note.created_by)
            
            note_data = [
                idx,
                note.title,
                note.content[:100] + '...' if len(note.content) > 100 else note.content,
                note.category,
                note.color,
                'نعم' if note.is_pinned else 'لا',
                creator.full_name if creator else 'مستخدم محذوف',
                note.created_at.strftime('%Y-%m-%d %H:%M') if note.created_at else 'غير محدد',
                note.updated_at.strftime('%Y-%m-%d %H:%M') if note.updated_at else 'غير محدد'
            ]
            
            for col, value in enumerate(note_data, 1):
                cell = ws_notes.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Instructor Notes data sheet
        ws_instructor_notes = wb.create_sheet(title="ملاحظات المدرسين")
        ws_instructor_notes.sheet_view.rightToLeft = True
        
        instructor_note_headers = ['#', 'العنوان', 'المحتوى', 'الطالب', 'المجموعة', 'الأولوية', 'الحالة', 'منشئ الملاحظة', 'مراجع من الإدارة', 'تاريخ الإنشاء', 'تاريخ المراجعة', 'رد الإدارة']
        for col, header in enumerate(instructor_note_headers, 1):
            cell = ws_instructor_notes.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        instructor_notes = InstructorNote.query.order_by(InstructorNote.created_at.desc()).all()
        for idx, note in enumerate(instructor_notes, 1):
            creator = User.query.get(note.created_by)
            reviewer = User.query.get(note.reviewed_by) if note.reviewed_by else None
            student = Student.query.get(note.student_id) if note.student_id else None
            group = Group.query.get(note.group_id) if note.group_id else None
            
            note_data = [
                idx,
                note.title,
                note.content[:100] + '...' if len(note.content) > 100 else note.content,
                student.name if student else 'غير محدد',
                group.name if group else 'غير محدد',
                note.priority,
                note.status,
                creator.full_name if creator else 'مستخدم محذوف',
                reviewer.full_name if reviewer else 'لم تتم المراجعة',
                note.created_at.strftime('%Y-%m-%d %H:%M') if note.created_at else 'غير محدد',
                note.reviewed_at.strftime('%Y-%m-%d %H:%M') if note.reviewed_at else 'لم تتم المراجعة',
                note.admin_response or 'لا يوجد رد'
            ]
            
            for col, value in enumerate(note_data, 1):
                cell = ws_instructor_notes.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Instructor Todos data sheet
        ws_instructor_todos = wb.create_sheet(title="مهام المدرسين")
        ws_instructor_todos.sheet_view.rightToLeft = True
        
        instructor_todo_headers = ['#', 'العنوان', 'الوصف', 'الفئة', 'الأولوية', 'الحالة', 'الطالب', 'المجموعة', 'تاريخ الاستحقاق', 'منشئ المهمة', 'تاريخ الإنشاء', 'تاريخ التحديث', 'تاريخ الإكمال']
        for col, header in enumerate(instructor_todo_headers, 1):
            cell = ws_instructor_todos.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        instructor_todos = InstructorTodo.query.order_by(InstructorTodo.created_at.desc()).all()
        for idx, todo in enumerate(instructor_todos, 1):
            creator = User.query.get(todo.created_by)
            student = Student.query.get(todo.student_id) if todo.student_id else None
            group = Group.query.get(todo.group_id) if todo.group_id else None
            
            todo_data = [
                idx,
                todo.title,
                todo.description[:100] + '...' if todo.description and len(todo.description) > 100 else todo.description or 'لا يوجد وصف',
                todo.category,
                todo.priority,
                todo.status,
                student.name if student else 'غير محدد',
                group.name if group else 'غير محدد',
                todo.due_date.strftime('%Y-%m-%d') if todo.due_date else 'غير محدد',
                creator.full_name if creator else 'مستخدم محذوف',
                todo.created_at.strftime('%Y-%m-%d %H:%M') if todo.created_at else 'غير محدد',
                todo.updated_at.strftime('%Y-%m-%d %H:%M') if todo.updated_at else 'غير محدد',
                todo.completed_at.strftime('%Y-%m-%d %H:%M') if todo.completed_at else 'غير مكتمل'
            ]
            
            for col, value in enumerate(todo_data, 1):
                cell = ws_instructor_todos.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Auto-adjust column widths for all sheets
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"نسخة_احتياطية_شاملة_نظام_تفرا_{timestamp}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'حدث خطأ أثناء إنشاء النسخة الاحتياطية: {str(e)}', 'error')
        return redirect(url_for('reports'))

@app.route('/import_system_data', methods=['GET', 'POST'])
@admin_required
def import_system_data():
    """Import complete system data from Excel file"""
    if request.method == 'GET':
        return render_template('import_data.html')
    
    try:
        if 'excel_file' not in request.files:
            flash('يرجى اختيار ملف Excel للاستيراد', 'error')
            return redirect(url_for('import_system_data'))
        
        file = request.files['excel_file']
        if file.filename == '':
            flash('يرجى اختيار ملف Excel للاستيراد', 'error')
            return redirect(url_for('import_system_data'))
        
        # Check if file is Excel
        if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            flash('يرجى رفع ملف Excel صحيح (.xlsx أو .xls)', 'error')
            return redirect(url_for('import_system_data'))
        
        # Read the Excel file
        from openpyxl import load_workbook
        import tempfile
        import os
        import time
        
        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            file.save(tmp_file.name)
            temp_file_path = tmp_file.name
            
        wb = None
        try:
            wb = load_workbook(temp_file_path, read_only=False, data_only=True)
            
            import_summary = {
                'users': 0, 'instructors': 0, 'students': 0, 'groups': 0,
                'schedules': 0, 'payments': 0, 'expenses': 0, 'errors': []
            }
            
            # Clear existing data if requested
            if request.form.get('clear_existing') == 'yes':
                # Clear all tables (except admin user)
                db.session.query(Attendance).delete()
                db.session.query(Payment).delete()
                db.session.query(Expense).delete()
                db.session.query(InstructorTodo).delete()
                db.session.query(InstructorNote).delete()
                db.session.query(Note).delete()
                db.session.query(Task).delete()
                db.session.query(Schedule).delete()
                
                # Clear many-to-many relationships
                db.session.execute(student_groups.delete())
                
                # Clear main entities
                db.session.query(Student).delete()
                db.session.query(Group).delete()
                db.session.query(Instructor).delete()
                
                # Keep only admin users
                db.session.query(User).filter(User.role != 'admin').delete()
                
                db.session.commit()
                flash('تم حذف البيانات الموجودة بنجاح', 'info')
            
            # Import Users (skip admin users to avoid conflicts)
            if 'المستخدمين' in wb.sheetnames:
                ws = wb['المستخدمين']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0] or not row[1]:  # Skip empty rows
                        continue
                    try:
                        username = str(row[1]).strip()
                        full_name = str(row[2]).strip()
                        role = str(row[3]).strip()
                        is_hidden = str(row[4]).strip() == 'نعم'
                        
                        # Skip admin users to avoid conflicts
                        if role == 'admin':
                            continue
                        
                        # Check if user already exists
                        if User.query.filter_by(username=username).first():
                            continue
                        
                        user = User(
                            username=username,
                            full_name=full_name,
                            role=role,
                            is_hidden=is_hidden
                        )
                        user.set_password('123456')  # Default password
                        db.session.add(user)
                        import_summary['users'] += 1
                    except Exception as e:
                        import_summary['errors'].append(f'خطأ في استيراد المستخدم {row}: {str(e)}')
                    
                    db.session.commit()
                
                # Import Instructors
                if 'المدرسين' in wb.sheetnames:
                    ws = wb['المدرسين']
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row[0] or not row[1]:  # Skip empty rows
                            continue
                        try:
                            name = str(row[1]).strip()
                            phone = str(row[2]).strip() if row[2] and str(row[2]).strip() != 'غير محدد' else None
                            specialization = str(row[3]).strip() if row[3] and str(row[3]).strip() != 'غير محدد' else None
                            
                            # Check if instructor already exists
                            if Instructor.query.filter_by(name=name).first():
                                continue
                            
                            instructor = Instructor(
                                name=name,
                                phone=phone,
                                specialization=specialization
                            )
                            db.session.add(instructor)
                            import_summary['instructors'] += 1
                        except Exception as e:
                            import_summary['errors'].append(f'خطأ في استيراد المدرس {row}: {str(e)}')
                    
                    db.session.commit()
                
                # Import Groups
                if 'المجموعات' in wb.sheetnames:
                    ws = wb['المجموعات']
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row[0] or not row[1]:  # Skip empty rows
                            continue
                        try:
                            name = str(row[1]).strip()
                            level = str(row[2]).strip() if row[2] and str(row[2]).strip() != 'غير محدد' else None
                            instructor_name = str(row[3]).strip() if row[3] and str(row[3]).strip() != 'غير محدد' else None
                            max_students = int(row[4]) if row[4] else 15
                            price = float(str(row[6]).replace(',', '')) if row[6] else 0.0
                            
                            # Find instructor
                            instructor = None
                            if instructor_name:
                                instructor = Instructor.query.filter_by(name=instructor_name).first()
                            
                            # Check if group already exists
                            if Group.query.filter_by(name=name).first():
                                continue
                            
                            group = Group(
                                name=name,
                                level=level,
                                instructor_id=instructor.id if instructor else None,
                                max_students=max_students,
                                price=price
                            )
                            db.session.add(group)
                            import_summary['groups'] += 1
                        except Exception as e:
                            import_summary['errors'].append(f'خطأ في استيراد المجموعة {row}: {str(e)}')
                    
                    db.session.commit()
                
                # Import Schedules
                if 'الجداول الزمنية' in wb.sheetnames:
                    ws = wb['الجداول الزمنية']
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row[0] or not row[1]:  # Skip empty rows
                            continue
                        try:
                            group_name = str(row[1]).strip()
                            day_of_week = str(row[3]).strip()
                            start_time = str(row[4]).strip()
                            end_time = str(row[5]).strip()
                            
                            # Find group
                            group = Group.query.filter_by(name=group_name).first()
                            if not group:
                                continue
                            
                            # Check if schedule already exists
                            if Schedule.query.filter_by(group_id=group.id, day_of_week=day_of_week).first():
                                continue
                            
                            schedule = Schedule(
                                group_id=group.id,
                                day_of_week=day_of_week,
                                start_time=start_time,
                                end_time=end_time
                            )
                            db.session.add(schedule)
                            import_summary['schedules'] += 1
                        except Exception as e:
                            import_summary['errors'].append(f'خطأ في استيراد الجدول الزمني {row}: {str(e)}')
                    
                    db.session.commit()
                
                # Import Students
                if 'الطلاب' in wb.sheetnames:
                    ws = wb['الطلاب']
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row[0] or not row[1]:  # Skip empty rows
                            continue
                        try:
                            name = str(row[1]).strip()
                            phone = str(row[2]).strip() if row[2] and str(row[2]).strip() != 'غير محدد' else None
                            age = int(row[3]) if row[3] and str(row[3]).strip() != 'غير محدد' else None
                            location = str(row[4]).strip() if row[4] and str(row[4]).strip() != 'غير محدد' else None
                            instructor_name = str(row[5]).strip() if row[5] and str(row[5]).strip() != 'غير محدد' else None
                            groups_names = str(row[6]).strip() if row[6] and str(row[6]).strip() != 'لا توجد مجموعات' else ''
                            
                            # Extract discount if available (column 8)
                            discount = 0.0
                            if len(row) > 8 and row[8]:
                                try:
                                    discount = float(str(row[8]).replace(',', ''))
                                except:
                                    discount = 0.0
                            
                            # Extract total_paid if available (column 10)
                            total_paid = 0.0
                            if len(row) > 10 and row[10]:
                                try:
                                    total_paid = float(str(row[10]).replace(',', ''))
                                except:
                                    total_paid = 0.0
                            
                            # Extract registration_date if available (column 12)
                            registration_date = datetime.now()
                            if len(row) > 12 and row[12]:
                                try:
                                    if isinstance(row[12], datetime):
                                        registration_date = row[12]
                                    else:
                                        registration_date = datetime.strptime(str(row[12]).split()[0], '%Y-%m-%d')
                                except:
                                    registration_date = datetime.now()
                            
                            # Find instructor
                            instructor = None
                            if instructor_name:
                                instructor = Instructor.query.filter_by(name=instructor_name).first()
                            
                            # Check if student already exists
                            if Student.query.filter_by(name=name).first():
                                continue
                            
                            student = Student(
                                name=name,
                                phone=phone,
                                age=age,
                                location=location,
                                instructor_id=instructor.id if instructor else None,
                                total_paid=total_paid,
                                discount=discount,
                                registration_date=registration_date
                            )
                            db.session.add(student)
                            db.session.flush()  # Get student ID
                            
                            # Add student to groups
                            if groups_names:
                                group_names_list = [g.strip() for g in groups_names.split(',')]
                                for group_name in group_names_list:
                                    group = Group.query.filter_by(name=group_name).first()
                                    if group:
                                        student.groups.append(group)
                            
                            import_summary['students'] += 1
                        except Exception as e:
                            import_summary['errors'].append(f'خطأ في استيراد الطالب {row}: {str(e)}')
                    
                    db.session.commit()
                
                # Import Payments
                if 'المدفوعات' in wb.sheetnames:
                    ws = wb['المدفوعات']
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row[0] or not row[1]:  # Skip empty rows
                            continue
                        try:
                            student_name = str(row[1]).strip()
                            amount = float(str(row[2]).replace(',', ''))
                            month = str(row[3]).strip() if row[3] and str(row[3]).strip() != 'غير محدد' else None
                            
                            # Parse date
                            payment_date = datetime.now()
                            if row[4]:
                                try:
                                    if isinstance(row[4], datetime):
                                        payment_date = row[4]
                                    else:
                                        payment_date = datetime.strptime(str(row[4]).split()[0], '%Y-%m-%d')
                                except:
                                    payment_date = datetime.now()
                            
                            notes = str(row[5]).strip() if row[5] and str(row[5]).strip() != 'لا توجد ملاحظات' else None
                            
                            # Find student
                            student = Student.query.filter_by(name=student_name).first()
                            if not student:
                                continue
                            
                            payment = Payment(
                                student_id=student.id,
                                amount=amount,
                                month=month,
                                date=payment_date,
                                notes=notes
                            )
                            db.session.add(payment)
                            import_summary['payments'] += 1
                        except Exception as e:
                            import_summary['errors'].append(f'خطأ في استيراد المدفوعة {row}: {str(e)}')
                    
                    db.session.commit()
                
                # Import Expenses
                if 'المصروفات' in wb.sheetnames:
                    ws = wb['المصروفات']
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row[0] or not row[1]:  # Skip empty rows
                            continue
                        try:
                            description = str(row[1]).strip()
                            amount = float(str(row[2]).replace(',', ''))
                            category = str(row[3]).strip() if row[3] and str(row[3]).strip() != 'غير محدد' else None
                            
                            # Parse date
                            expense_date = datetime.now()
                            if row[4]:
                                try:
                                    if isinstance(row[4], datetime):
                                        expense_date = row[4]
                                    else:
                                        expense_date = datetime.strptime(str(row[4]).split()[0], '%Y-%m-%d')
                                except:
                                    expense_date = datetime.now()
                            
                            notes = str(row[5]).strip() if row[5] and str(row[5]).strip() != 'لا توجد ملاحظات' else None
                            
                            expense = Expense(
                                description=description,
                                amount=amount,
                                category=category,
                                date=expense_date,
                                notes=notes
                            )
                            db.session.add(expense)
                            import_summary['expenses'] += 1
                        except Exception as e:
                            import_summary['errors'].append(f'خطأ في استيراد المصروف {row}: {str(e)}')
                    
                    db.session.commit()
                
                # Generate success message
                success_msg = f"تم استيراد البيانات بنجاح! "
                success_msg += f"المستخدمين: {import_summary['users']}, "
                success_msg += f"المدرسين: {import_summary['instructors']}, "
                success_msg += f"الطلاب: {import_summary['students']}, "
                success_msg += f"المجموعات: {import_summary['groups']}, "
                success_msg += f"الجداول: {import_summary['schedules']}, "
                success_msg += f"المدفوعات: {import_summary['payments']}, "
                success_msg += f"المصروفات: {import_summary['expenses']}"
                
                flash(success_msg, 'success')
                
                # Show errors if any
                if import_summary['errors']:
                    for error in import_summary['errors'][:5]:  # Show first 5 errors
                        flash(error, 'warning')
                    if len(import_summary['errors']) > 5:
                        flash(f'وتوجد {len(import_summary["errors"]) - 5} أخطاء أخرى...', 'warning')
                
                return redirect(url_for('reports'))
                
        finally:
                # Close workbook properly to release file handle
                if wb:
                    try:
                        wb.close()
                    except:
                        pass
                
                # Clean up temporary file with multiple attempts
                cleanup_attempts = 0
                max_attempts = 3
                while cleanup_attempts < max_attempts:
                    try:
                        os.unlink(temp_file_path)
                        break  # Success - exit loop
                    except Exception as cleanup_error:
                        cleanup_attempts += 1
                        if cleanup_attempts < max_attempts:
                            # Wait and try again
                            time.sleep(0.1)
                        else:
                            # Log warning for last attempt
                            print(f"Warning: Could not delete temporary file {temp_file_path} after {max_attempts} attempts: {cleanup_error}")
                            # File will be cleaned up by system later
                            break
                
    except Exception as e:
        flash(f'حدث خطأ أثناء استيراد البيانات: {str(e)}', 'error')
        return redirect(url_for('import_system_data'))

@app.route('/admin_respond_instructor_note/<int:note_id>', methods=['POST'])
@admin_required
def admin_respond_instructor_note(note_id):
    """Admin responds to instructor note"""
    try:
        current_user = get_current_user()
        note = InstructorNote.query.get_or_404(note_id)
        
        response = request.form.get('response', '')
        status = request.form.get('status', 'قيد المراجعة')
        
        note.admin_response = response
        note.status = status
        note.reviewed_by = current_user.id
        note.reviewed_at = datetime.utcnow()
        
        db.session.commit()
        flash('تم الرد على الملاحظة بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash('حدث خطأ أثناء الرد على الملاحظة', 'error')
    
    return redirect(url_for('tasks') + '#instructor-notes')

@app.route('/admin_update_instructor_note_status/<int:note_id>', methods=['POST'])
@admin_required
def admin_update_instructor_note_status(note_id):
    """Admin updates instructor note status"""
    try:
        current_user = get_current_user()
        note = InstructorNote.query.get_or_404(note_id)
        
        new_status = request.form['status']
        note.status = new_status
        
        # If not reviewed yet, mark as reviewed
        if not note.reviewed_by:
            note.reviewed_by = current_user.id
            note.reviewed_at = datetime.utcnow()
        
        db.session.commit()
        flash('تم تحديث حالة الملاحظة بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash('حدث خطأ أثناء تحديث حالة الملاحظة', 'error')
    
    return redirect(url_for('tasks') + '#instructor-notes')

if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5000))
    # Enable debug mode for development by default
    debug = os.environ.get('FLASK_ENV') != 'production'
    app.run(host='0.0.0.0', port=port, debug=True)
else:
    # Production mode: Initialize database when imported
    init_db() 